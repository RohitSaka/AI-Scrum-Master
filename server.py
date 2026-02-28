from fastapi import FastAPI, HTTPException, Header, Depends, BackgroundTasks, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, RedirectResponse, FileResponse
import requests, json, os, uuid, time, traceback, math
from requests.auth import HTTPBasicAuth
from datetime import datetime, timedelta
from dotenv import load_dotenv
import urllib.parse
import io
import base64
import csv, io, math, calendar
from pydantic import BaseModel
from typing import List, Optional
from strategic_roadmap import assemble_strategic_roadmap, DEFAULT_INVESTMENT_BUCKETS
# --- NATIVE PPTX GENERATION (v5 Engine) ---
from pptx_engine_v5 import generate_native_editable_pptx, THEMES

# --- MEETING AGENT ---
from meeting_agent import (
    process_meeting_transcript, classify_meeting, fetch_sprint_history,
    calculate_velocity, generate_capacity_report
)

# --- DATABASE (SQLAlchemy) ---
from sqlalchemy import create_engine, Column, String, Boolean, DateTime, Integer, Text
from sqlalchemy.orm import declarative_base, sessionmaker, Session

load_dotenv()
app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

print("\n" + "="*60)
print("\U0001f680 APP STARTING: V49 \u2014 MEETING AGENT + PPTX v6")
print("   Meeting Agent | Auto Stories | Capacity Planning | 4 Themes")
print("   Sprint=DarkTeal | Weekly=LightCorp | Monthly=Executive | Quarterly=PremiumDark")
print("="*60 + "\n")

# ================= DATABASE SETUP =================
DATABASE_URL = os.getenv("DATABASE_URL", "sqlite:///./local_agile.db")
if DATABASE_URL.startswith("postgres://"):
    DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql://", 1)

engine_db = create_engine(DATABASE_URL, connect_args={"check_same_thread": False} if "sqlite" in DATABASE_URL else {})
SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine_db)
Base = declarative_base()

class License(Base):
    __tablename__ = "licenses"
    key = Column(String, primary_key=True, index=True)
    is_active = Column(Boolean, default=True)
    created_at = Column(DateTime, default=datetime.utcnow)

class UserAuth(Base):
    __tablename__ = "user_auth"
    id = Column(Integer, primary_key=True, index=True)
    license_key = Column(String, unique=True, index=True)
    access_token = Column(Text)
    refresh_token = Column(Text)
    cloud_id = Column(String)
    expires_at = Column(Integer)

class GuestLink(Base):
    __tablename__ = "guest_links"
    token = Column(String, primary_key=True, index=True)
    project_key = Column(String)
    sprint_id = Column(String)
    license_key = Column(String)
    created_at = Column(DateTime, default=datetime.utcnow)

class MeetingSession(Base):
    __tablename__ = "meeting_sessions"
    id = Column(String, primary_key=True, index=True)
    license_key = Column(String)
    project_key = Column(String)
    sprint_id = Column(String)
    meeting_type = Column(String)       # planning/grooming/retro/capacity
    transcript = Column(Text)
    ai_results = Column(Text)           # JSON of extracted items
    status = Column(String, default="pending")
    created_at = Column(DateTime, default=datetime.utcnow)
    platform = Column(String)           # google_meet/teams/zoom/manual

Base.metadata.create_all(bind=engine_db)

class FeatureRoadmapRequest(BaseModel):
    tech_stack: str
    features: List[str]
    target_duration_value: float         # was int — now accepts 1.5, 2.5 etc.
    target_duration_unit: str
    project_key: Optional[str] = None
    start_date: Optional[str] = None  # ISO date string, defaults to today

class JiraPushRequest(BaseModel):
    epics: List[dict]

class XLSXDownloadRequest(BaseModel):
    data: dict

def get_db():
    db = SessionLocal()
    try: yield db
    finally: db.close()

# ================= OAUTH 2.0 & LICENSING =================
CLIENT_ID = os.getenv("ATLASSIAN_CLIENT_ID", "").strip()
CLIENT_SECRET = os.getenv("ATLASSIAN_CLIENT_SECRET", "").strip()
APP_URL = os.getenv("APP_URL", "http://localhost:8000").strip()
REDIRECT_URI = f"{APP_URL}/auth/callback"

@app.post("/admin/generate_license")
def generate_license(db: Session = Depends(get_db)):
    new_key = f"IG-ENT-{str(uuid.uuid4())[:8].upper()}"
    db.add(License(key=new_key))
    db.commit()
    return {"license_key": new_key, "status": "active"}

@app.get("/auth/login")
def login(license_key: str, db: Session = Depends(get_db)):
    lic = db.query(License).filter(License.key == license_key).first()
    if not lic or not lic.is_active: raise HTTPException(status_code=403, detail="Invalid License Key")
    params = {"audience": "api.atlassian.com", "client_id": CLIENT_ID, "scope": "read:jira-work manage:jira-project manage:jira-configuration write:jira-work offline_access", "redirect_uri": REDIRECT_URI, "state": license_key, "response_type": "code", "prompt": "consent"}
    return RedirectResponse(f"https://auth.atlassian.com/authorize?{urllib.parse.urlencode(params, quote_via=urllib.parse.quote)}")

@app.get("/auth/callback")
def auth_callback(code: str, state: str, db: Session = Depends(get_db)):
    license_key = state
    res = requests.post("https://auth.atlassian.com/oauth/token", json={"grant_type": "authorization_code", "client_id": CLIENT_ID, "client_secret": CLIENT_SECRET, "code": code, "redirect_uri": REDIRECT_URI}, timeout=30)
    if res.status_code != 200: raise HTTPException(status_code=400, detail="OAuth Failed")
    tokens = res.json()
    cloud_id = requests.get("https://api.atlassian.com/oauth/token/accessible-resources", headers={"Authorization": f"Bearer {tokens['access_token']}"}, timeout=30).json()[0]["id"]
    user = db.query(UserAuth).filter(UserAuth.license_key == license_key).first()
    if not user: user = UserAuth(license_key=license_key); db.add(user)
    user.access_token = tokens.get("access_token"); user.refresh_token = tokens.get("refresh_token", ""); user.expires_at = int(time.time()) + tokens.get("expires_in", 3600); user.cloud_id = cloud_id
    db.commit()
    requests.post(f"https://api.atlassian.com/ex/jira/{cloud_id}/rest/api/3/webhook", headers={"Authorization": f"Bearer {tokens.get('access_token')}", "Content-Type": "application/json"}, json={"url": f"{APP_URL}/webhook?cloud_id={cloud_id}", "webhooks": [{"events": ["jira:issue_created"], "jqlFilter": "project IS NOT EMPTY"}]}, timeout=30)
    return RedirectResponse(f"{APP_URL}/?success=true")

def get_valid_oauth_session(db: Session, license_key: str = None, cloud_id: str = None):
    if license_key: user = db.query(UserAuth).filter(UserAuth.license_key == license_key).first()
    elif cloud_id: user = db.query(UserAuth).filter(UserAuth.cloud_id == cloud_id).first()
    else: user = db.query(UserAuth).order_by(UserAuth.expires_at.desc()).first()
    if not user: return None
    if int(time.time()) >= user.expires_at - 300:
        res = requests.post("https://auth.atlassian.com/oauth/token", json={"grant_type": "refresh_token", "client_id": CLIENT_ID, "client_secret": CLIENT_SECRET, "refresh_token": user.refresh_token}, timeout=30)
        if res.status_code == 200:
            tokens = res.json(); user.access_token = tokens.get("access_token")
            if tokens.get("refresh_token"): user.refresh_token = tokens.get("refresh_token")
            user.expires_at = int(time.time()) + tokens.get("expires_in", 3600); db.commit()
    return user

async def get_jira_creds(x_jira_domain: str = Header(None), x_jira_email: str = Header(None), x_jira_token: str = Header(None), x_license_key: str = Header(None), db: Session = Depends(get_db)):
    if x_license_key:
        user = get_valid_oauth_session(db=db, license_key=x_license_key)
        if not user: raise HTTPException(status_code=401, detail="Invalid License")
        return {"auth_type": "oauth", "cloud_id": user.cloud_id, "access_token": user.access_token}
    if x_jira_domain and x_jira_email and x_jira_token:
        return {"auth_type": "basic", "domain": x_jira_domain.replace("https://", "").replace("http://", "").strip("/"), "email": x_jira_email, "token": x_jira_token}
    raise HTTPException(status_code=401, detail="Missing Auth")

def jira_request(method, endpoint, creds, data=None):
    try:
        if creds.get("auth_type") == "oauth":
            url = f"https://api.atlassian.com/ex/jira/{creds.get('cloud_id')}/rest/api/3/{endpoint}"
            headers = {"Accept": "application/json", "Content-Type": "application/json", "Authorization": f"Bearer {creds.get('access_token')}"}
            auth = None
        else:
            url = f"https://{creds['domain']}/rest/api/3/{endpoint}"
            headers = {"Accept": "application/json", "Content-Type": "application/json"}
            auth = HTTPBasicAuth(creds['email'], creds['token'])
        if method == "POST": return requests.post(url, json=data, headers=headers, auth=auth, timeout=60)
        elif method == "GET": return requests.get(url, headers=headers, auth=auth, timeout=60)
        elif method == "PUT": return requests.put(url, json=data, headers=headers, auth=auth, timeout=60)
    except Exception as e:
        print(f"Jira HTTP Error ({endpoint}): {e}", flush=True)
        return None

# ================= JIRA LOGIC & AI CORE =================
STORY_POINT_CACHE = {}

def get_assignable_users(project_key, creds):
    res = jira_request("GET", f"user/assignable/search?project={project_key}", creds)
    users = {}
    if res is not None and res.status_code == 200:
        for u in res.json():
            if 'displayName' in u and 'accountId' in u: users[u['displayName']] = u['accountId']
    return users

def build_team_roster(project_key, creds, sp_field):
    assignable_map = get_assignable_users(project_key, creds)
    roster = {name: 0.0 for name in assignable_map.keys()}
    res = jira_request("POST", "search/jql", creds, {"jql": f'project="{project_key}" AND sprint in openSprints()', "fields": ["assignee", sp_field]})
    if res is not None and res.status_code == 200:
        for i in res.json().get('issues', []):
            f = i.get('fields') or {}
            name = (f.get('assignee') or {}).get('displayName')
            if name and name in roster: roster[name] += extract_story_points(f, sp_field)
    return roster, assignable_map

def get_story_point_field(creds):
    domain_key = creds.get('domain') or creds.get('cloud_id')
    if domain_key in STORY_POINT_CACHE: return STORY_POINT_CACHE[domain_key]
    res = jira_request("GET", "field", creds)
    if res is not None and res.status_code == 200:
        fields = res.json()
        for f in fields:
            if f.get('name', '').lower() == "story point estimate": STORY_POINT_CACHE[domain_key] = f['id']; return f['id']
        for f in fields:
            if "story point" in f.get('name', '').lower() or "estimate" in f.get('name', '').lower(): STORY_POINT_CACHE[domain_key] = f['id']; return f['id']
    return "customfield_10016"

def safe_float(val):
    try: return float(val) if val is not None else 0.0
    except: return 0.0

def extract_story_points(issue_fields, sp_field):
    pts = safe_float(issue_fields.get(sp_field))
    if pts > 0: return pts
    for field in ['customfield_10016', 'customfield_10026', 'customfield_10004', 'customfield_10028']:
        val = safe_float(issue_fields.get(field))
        if val > 0: return val
    for key, value in issue_fields.items():
        if key.startswith("customfield_") and isinstance(value, (int, float)) and 0 < value < 100: return float(value)
    return 0.0

def get_jira_account_id(display_name, creds):
    if not display_name or display_name.lower() in ["unassigned", "none", "null"]: return None
    safe_query = urllib.parse.quote(display_name)
    res = jira_request("GET", f"user/search?query={safe_query}", creds)
    if res is not None and res.status_code == 200 and res.json():
        users = res.json()
        if isinstance(users, list) and len(users) > 0: return users[0].get("accountId")
    return None

def extract_adf_text(adf_node):
    if not adf_node or not isinstance(adf_node, dict): return ""
    text = ""
    if adf_node.get('type') == 'text': text += adf_node.get('text', '') + " "
    for content in adf_node.get('content', []): text += extract_adf_text(content)
    return text.strip()

def extract_jira_error(res):
    if res is None: return "Connection Timeout or Network Error."
    try:
        data = res.json(); errs = []
        if "errorMessages" in data: errs.extend(data["errorMessages"])
        if "errors" in data:
            for k, v in data["errors"].items(): errs.append(f"{k}: {v}")
        if errs: return " | ".join(errs)
    except: pass
    return f"Status {res.status_code}: {res.text[:200]}"

def create_adf_doc(text_content, ac_list=None):
    blocks = []
    for line in str(text_content).split('\n'):
        clean_line = line.strip()
        if clean_line: blocks.append({"type": "paragraph", "content": [{"type": "text", "text": clean_line}]})
    if ac_list and isinstance(ac_list, list):
        blocks.append({"type": "heading", "attrs": {"level": 3}, "content": [{"type": "text", "text": "Acceptance Criteria"}]})
        list_items = [{"type": "listItem", "content": [{"type": "paragraph", "content": [{"type": "text", "text": str(ac)}]}]} for ac in ac_list if str(ac).strip()]
        if list_items: blocks.append({"type": "bulletList", "content": list_items})
    if not blocks: blocks.append({"type": "paragraph", "content": [{"type": "text", "text": "AI Generated Content"}]})
    return {"type": "doc", "version": 1, "content": blocks}

def call_gemini(prompt, temperature=0.3, image_data=None, json_mode=True, timeout=50, model=None):
    """Call Gemini via Vertex AI (Enterprise) using google-genai SDK."""
    from google import genai
    from google.oauth2 import service_account as sa

    # ── Initialize Vertex AI client (cached after first call) ──
    if not hasattr(call_gemini, "_client"):
        creds_json_str = os.getenv("GOOGLE_CREDENTIALS_JSON", "")
        project_id = os.getenv("GOOGLE_CLOUD_PROJECT", "")
        api_key = os.getenv("GEMINI_API_KEY", "")

        if creds_json_str and project_id:
            # Vertex AI Enterprise path
            creds = sa.Credentials.from_service_account_info(
                json.loads(creds_json_str),
                scopes=["https://www.googleapis.com/auth/cloud-platform"]
            )
            call_gemini._client = genai.Client(
                vertexai=True,
                project=project_id,
                location=os.getenv("VERTEX_AI_LOCATION", "us-central1"),
                credentials=creds,
            )
            call_gemini._mode = "vertex"
            print("✅ Vertex AI Enterprise client initialized", flush=True)
        elif api_key:
            # Fallback: AI Studio (API key)
            call_gemini._client = genai.Client(api_key=api_key)
            call_gemini._mode = "aistudio"
            print("⚠️ Using AI Studio (API key) — set GOOGLE_CREDENTIALS_JSON for Vertex AI", flush=True)
        else:
            call_gemini._client = None
            call_gemini._mode = "none"
            print("❌ No Gemini credentials found", flush=True)

    client = call_gemini._client
    if not client:
        return None

    # ── Build content parts ──
    parts = [prompt]
    if image_data:
        try:
            header, encoded = image_data.split(",", 1)
            mime_type = header.split(":")[1].split(";")[0]
            import base64 as b64
            parts.append(genai.types.Part.from_bytes(data=b64.b64decode(encoded), mime_type=mime_type))
        except Exception as e:
            print(f"Image Parse Error: {e}", flush=True)

    if model:
        models_to_try = [model, "gemini-3.1-pro"]
    else:
        models_to_try = ["gemini-2.5-flash", "gemini-2.5-flash-lite"]

    for m in models_to_try:
        try:
            config = {"temperature": temperature, "max_output_tokens": 8192}
            if json_mode:
                config["response_mime_type"] = "application/json"

            response = client.models.generate_content(
                model=m,
                contents=parts,
                config=config,
            )
            if response and response.text:
                return response.text
        except Exception as e:
            err = str(e).lower()
            if "429" in err or "quota" in err or "rate" in err or "resource_exhausted" in err:
                print(f"Gemini rate limit on {m}, trying next model...", flush=True)
                import time as _t; _t.sleep(2)
                continue
            print(f"Gemini {m} error: {e}", flush=True)
            continue
    return None

def call_openai(prompt, temperature=0.3, image_data=None, json_mode=True, timeout=20, model="gpt-4o"):
    api_key = os.getenv("OPENAI_API_KEY")
    if not api_key: return call_gemini(prompt, temperature, image_data, json_mode, timeout=timeout)
    sys_msg = "You are an elite Enterprise Strategy Consultant. Return strictly valid JSON." if json_mode else "You are an Expert Agile Coach assisting a Scrum Master."
    messages = [{"role": "system", "content": sys_msg}, {"role": "user", "content": [{"type": "text", "text": prompt}]}]
    if image_data: messages[1]["content"].append({"type": "image_url", "image_url": {"url": image_data}})
    try:
        kwargs = {"model": model, "messages": messages, "temperature": temperature}
        if json_mode: kwargs["response_format"] = {"type": "json_object"}
        r = requests.post("https://api.openai.com/v1/chat/completions", headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"}, json=kwargs, timeout=timeout)
        if r.status_code == 200: return r.json()['choices'][0]['message']['content']
    except Exception: pass
    print("Seamless Fallback to Google Gemini...", flush=True)
    return call_gemini(prompt, temperature, image_data, json_mode, timeout=timeout)


def generate_ai_response(prompt, temperature=0.3, force_openai=False, image_data=None, json_mode=True, timeout=20, model=None):
    """
    Routes AI calls. Default: Gemini. 
    force_openai=True only for endpoints that specifically need OpenAI (e.g. image analysis).
    model param: pass specific Gemini model like "gemini-3.1-pro" for premium tasks.
    """
    if force_openai:
        return call_openai(prompt, temperature, image_data, json_mode, timeout=timeout, model=model or "gpt-4o")
    return call_gemini(prompt, temperature, image_data, json_mode, timeout=timeout, model=model)

# ================= APP ENDPOINTS =================
@app.get("/")
def home():
    if os.path.exists("index.html"): return FileResponse("index.html")
    return {"status": "Backend running — V48 PPTX Engine v5 + Roles & Team"}

@app.get("/projects")
def list_projects(creds: dict = Depends(get_jira_creds)):
    res = jira_request("GET", "project", creds)
    try: return [{"key": p["key"], "name": p["name"], "avatar": p["avatarUrls"]["48x48"]} for p in res.json()]
    except: return []

@app.get("/sprints/{project_key}")
def get_sprints(project_key: str, creds: dict = Depends(get_jira_creds)):
    res = jira_request("POST", "search/jql", creds, {"jql": f'project="{project_key}" AND sprint is not EMPTY ORDER BY updated DESC', "maxResults": 50, "fields": ["customfield_10020"]})
    try:
        sprints = {}
        for i in res.json().get('issues', []):
            for s in (i.get('fields') or {}).get('customfield_10020') or []: sprints[s['id']] = {"id": s['id'], "name": s['name'], "state": s['state']}
        return sorted(list(sprints.values()), key=lambda x: x['id'], reverse=True)
    except: return []

@app.get("/analytics/{project_key}")
def get_analytics(project_key: str, sprint_id: str = None, creds: dict = Depends(get_jira_creds)):
    sp_field = get_story_point_field(creds)
    jql = f'project="{project_key}" AND sprint={sprint_id}' if sprint_id and sprint_id != "active" else f'project="{project_key}" AND sprint in openSprints()'
    safe_fields = ["summary", "assignee", "priority", "status", "issuetype", "description", "comment", "created", "customfield_10020", "customfield_10016", "customfield_10026", "customfield_10028", "customfield_10004", sp_field]
    res = jira_request("POST", "search/jql", creds, {"jql": jql, "maxResults": 100, "fields": safe_fields})
    if res is None or res.status_code != 200:
        return {"metrics": {"total": 0, "points": 0.0, "blockers": 0, "bugs": 0, "stories": 0, "assignees": {}}, "ai_insights": {}}
    issues = res.json().get('issues', [])
    stats = {"total": len(issues), "points": 0.0, "blockers": 0, "bugs": 0, "stories": 0, "assignees": {}}
    context_for_ai = []
    for i in issues:
        f = i.get('fields') or {}
        assignee = f.get('assignee') or {}; priority = f.get('priority') or {}; status = f.get('status') or {}; issuetype = f.get('issuetype') or {}
        name = assignee.get('displayName') or "Unassigned"; pts = extract_story_points(f, sp_field); priority_name = priority.get('name') or "Medium"; status_name = status.get('name') or "To Do"
        added_mid_sprint = False
        try:
            sprint_start = None
            for k, v in f.items():
                if isinstance(v, list) and len(v) > 0 and isinstance(v[0], dict) and 'startDate' in v[0] and 'state' in v[0]:
                    for s in v:
                        if str(s.get('state', '')).lower() == 'active' or str(s.get('id', '')) == str(sprint_id): sprint_start = s.get('startDate')
                    if sprint_start: break
            if sprint_start:
                created_date = f.get('created', '')
                if created_date and str(created_date) > str(sprint_start): added_mid_sprint = True
        except Exception: pass
        stats["points"] += pts
        if priority_name in ["High", "Highest", "Critical"]: stats["blockers"] += 1
        if issuetype.get('name') == "Bug": stats["bugs"] += 1
        if name not in stats["assignees"]: stats["assignees"][name] = {"count": 0, "points": 0.0, "tasks": [], "avatar": assignee.get('avatarUrls', {}).get('48x48', '')}
        stats["assignees"][name]["count"] += 1; stats["assignees"][name]["points"] += pts
        stats["assignees"][name]["tasks"].append({"key": i.get('key'), "summary": f.get('summary', ''), "points": pts, "status": status_name, "priority": priority_name, "added_mid_sprint": added_mid_sprint})
        desc = extract_adf_text(f.get('description', {}))[:500]
        context_for_ai.append({"key": i.get('key'), "status": status_name, "assignee": name, "summary": f.get('summary', ''), "description": desc})
    try:
        raw_ai = generate_ai_response(f"Analyze Sprint. DATA: {json.dumps(context_for_ai)}. Return JSON: {{\"executive_summary\": \"...\", \"business_value\": \"...\", \"story_progress\": [{{\"key\":\"...\", \"summary\":\"...\", \"assignee\":\"...\", \"status\":\"...\", \"analysis\":\"...\"}}]}}").replace('```json','').replace('```','').strip()
        ai_data = json.loads(raw_ai)
        if "executive_summary" not in ai_data: raise ValueError("Bad format")
    except Exception as e:
        ai_data = {"executive_summary": "Format Error.", "business_value": "Error", "story_progress": []}
    return {"metrics": stats, "ai_insights": ai_data}

@app.get("/super_deck/{project_key}")
def generate_super_deck(project_key: str, sprint_id: str = None, creds: dict = Depends(get_jira_creds)):
    """Generate Sprint Review deck with SPRINT theme."""
    sp_field = get_story_point_field(creds)
    jql = f'project="{project_key}" AND sprint={sprint_id}' if sprint_id and sprint_id != "active" else f'project="{project_key}" AND sprint in openSprints()'
    safe_fields = ["summary", "status", "priority", "assignee", "customfield_10016", "customfield_10026", "customfield_10028", "customfield_10004", sp_field]
    res = jira_request("POST", "search/jql", creds, {"jql": jql, "maxResults": 30, "fields": safe_fields})
    issues = res.json().get('issues', []) if res is not None and res.status_code == 200 else []
    done_pts = 0.0; total_pts = 0.0; active_users = set(); blockers = []; done_summaries = []
    for i in issues:
        f = i.get('fields') or {}; status_category = (f.get('status') or {}).get('statusCategory') or {}; priority = f.get('priority') or {}; assignee = f.get('assignee') or {}
        pts = extract_story_points(f, sp_field); total_pts += pts
        if status_category.get('key') == 'done': done_pts += pts; done_summaries.append(f.get('summary', ''))
        if priority.get('name') in ["High", "Highest", "Critical"]: blockers.append(f.get('summary', ''))
        if assignee: active_users.add(assignee.get('displayName', ''))
    retro_res = jira_request("GET", f"project/{project_key}/properties/ig_agile_retro", creds)
    retro_data = retro_res.json().get('value', {}).get(str(sprint_id) if sprint_id else 'active', {}) if retro_res is not None and retro_res.status_code==200 else {}
    backlog_res = jira_request("POST", "search/jql", creds, {"jql": f'project="{project_key}" AND sprint is EMPTY', "maxResults": 4, "fields": ["summary"]})
    backlog = [i.get('fields', {}).get('summary') for i in backlog_res.json().get('issues', [])] if backlog_res is not None and backlog_res.status_code == 200 else ["Backlog Refinement", "Planning"]
    context = {"project": project_key, "current_date": datetime.now().strftime("%B %d, %Y"), "total_points": total_pts, "completed_points": done_pts, "blockers": blockers[:3], "retro": retro_data, "accomplishments": done_summaries[:4], "backlog_preview": backlog}
    prompt = f"""Act as a McKinsey Agile Consultant. Build a 6-Slide Sprint Report based on this exact data: {json.dumps(context)}.
CRITICAL INSTRUCTION: DO NOT USE PLACEHOLDERS. WRITE FULL PROFESSIONAL SENTENCES FROM THE REAL DATA.
IMPORTANT: For the flowchart slide (slide 6), the "items" field MUST be an array of objects with "title" keys like: [{{"title": "Backlog item name"}}, {{"title": "Another item"}}]. Use real backlog items from the data.
Return EXACTLY a JSON array:
[
  {{"id": 1, "layout": "hero", "title": "Sprint Review", "subtitle": "{context['current_date']}"}},
  {{"id": 2, "layout": "standard", "title": "Executive Summary", "content": ["Real sentence 1", "Real sentence 2", "Real sentence 3"]}},
  {{"id": 3, "layout": "kpi_grid", "title": "Sprint Metrics", "items": [{{"label": "Velocity Delivered", "value": "{done_pts}"}}, {{"label": "Total Planned", "value": "{total_pts}"}}, {{"label": "Team Members", "value": "{len(active_users)}"}}]}},
  {{"id": 4, "layout": "icon_columns", "title": "Risks & Blockers", "items": [{{"title": "Blocker Title", "text": "Real description from data", "icon": ""}}]}},
  {{"id": 5, "layout": "standard", "title": "Continuous Improvement", "content": ["Real retro insights from data"]}},
  {{"id": 6, "layout": "flowchart", "title": "Next Sprint Plan", "items": [{{"title": "Real backlog item 1"}}, {{"title": "Real backlog item 2"}}, {{"title": "Real backlog item 3"}}]}}
]"""
    try:
        raw = generate_ai_response(prompt, temperature=0.5, force_openai=True).replace('```json','').replace('```','').strip()
        return {"status": "success", "slides": json.loads(raw), "theme": "sprint"}
    except Exception as e:
        print(f"Deck Parse Error: {e}", flush=True)
        return {"status": "error", "message": "Failed to orchestrate slides."}

@app.get("/report_deck/{project_key}/{timeframe}")
def generate_report_deck(project_key: str, timeframe: str, creds: dict = Depends(get_jira_creds)):
    """Generate timeframe-specific report deck with DISTINCT theme per timeframe."""
    sp_field = get_story_point_field(creds)
    days = 7 if timeframe == "weekly" else (30 if timeframe == "monthly" else 90)
    dt = (datetime.now() - timedelta(days=days)).strftime("%Y-%m-%d")
    safe_fields = ["summary", "status", "assignee", "priority", "customfield_10016", "customfield_10026", "customfield_10028", "customfield_10004", sp_field]
    res = jira_request("POST", "search/jql", creds, {"jql": f'project="{project_key}" AND updated >= \"{dt}\" ORDER BY updated DESC', "maxResults": 40, "fields": safe_fields})
    issues = res.json().get('issues', []) if res is not None and res.status_code == 200 else []
    done_count = 0; done_pts = 0.0; accomplishments = []; blockers = []
    for i in issues:
        f = i.get('fields') or {}; pts = extract_story_points(f, sp_field)
        if (f.get('status') or {}).get('statusCategory', {}).get('key') == 'done': done_count += 1; done_pts += pts; accomplishments.append(f.get('summary', ''))
        if (f.get('priority') or {}).get('name') in ["High", "Highest", "Critical"]: blockers.append(f.get('summary', ''))
    context = {"project": project_key, "timeframe": timeframe.capitalize(), "current_date": datetime.now().strftime("%B %d, %Y"), "completed_issues": done_count, "completed_velocity": done_pts, "accomplishments": accomplishments[:5], "blockers": blockers[:3]}

    # ═══ CRITICAL: Map each timeframe to its DISTINCT theme ═══
    theme_map = {"weekly": "weekly", "monthly": "monthly", "quarterly": "quarterly"}
    theme_name = theme_map.get(timeframe, "sprint")

    agendas = {
        "weekly": f"""[
  {{"layout": "hero", "title": "Weekly Status Report", "subtitle": "{context['current_date']}"}},
  {{"layout": "kpi_grid", "title": "Weekly Metrics", "items": [{{"label": "Issues Closed", "value": "{done_count}"}}, {{"label": "Points Delivered", "value": "{done_pts}"}}]}},
  {{"layout": "standard", "title": "Key Accomplishments", "content": ["Real bullet 1", "Real bullet 2"]}},
  {{"layout": "icon_columns", "title": "Risks & Action Items", "items": [{{"title": "Risk", "text": "Real description"}}]}},
  {{"layout": "flowchart", "title": "Next Week Priorities", "items": [{{"title": "Priority 1"}}, {{"title": "Priority 2"}}, {{"title": "Priority 3"}}]}}
]""",
        "monthly": f"""[
  {{"layout": "hero", "title": "Monthly Business Review", "subtitle": "{context['current_date']}"}},
  {{"layout": "standard", "title": "Executive Summary", "content": ["Real bullet 1", "Real bullet 2"]}},
  {{"layout": "kpi_grid", "title": "Key Performance Indicators", "items": [{{"label": "Velocity", "value": "{done_pts}"}}, {{"label": "Completion Rate", "value": "{done_count}"}}]}},
  {{"layout": "icon_columns", "title": "Strategic Wins", "items": [{{"title": "Win 1", "text": "Details"}}]}},
  {{"layout": "standard", "title": "Risks & Mitigation", "content": ["Real bullet 1", "Real bullet 2"]}},
  {{"layout": "flowchart", "title": "Next Month Initiatives", "items": [{{"title": "Goal 1"}}, {{"title": "Goal 2"}}]}}
]""",
        "quarterly": f"""[
  {{"layout": "hero", "title": "Quarterly Business Review", "subtitle": "{context['current_date']}"}},
  {{"layout": "standard", "title": "Quarter in Review", "content": ["Real bullet 1", "Real bullet 2"]}},
  {{"layout": "icon_columns", "title": "Business Impact", "items": [{{"title": "Impact 1", "text": "Details"}}]}},
  {{"layout": "kpi_grid", "title": "Quarterly Metrics", "items": [{{"label": "Total Velocity", "value": "{done_pts}"}}, {{"label": "Issues Resolved", "value": "{done_count}"}}]}},
  {{"layout": "flowchart", "title": "Strategic Roadmap", "items": [{{"title": "Milestone 1"}}, {{"title": "Milestone 2"}}, {{"title": "Milestone 3"}}]}}
]"""
    }

    prompt = f"""Act as an Elite Enterprise Designer. Create a {timeframe.capitalize()} Business Review Deck for project {project_key} based ONLY on this data: {json.dumps(context)}.
CRITICAL: WRITE REAL TEXT FROM THE DATA. DO NOT OUTPUT PLACEHOLDERS.
IMPORTANT: For all flowchart slides, the "items" field MUST be an array of objects like: [{{"title": "Item name"}}]. Always include at least 3 items.
Return EXACTLY a JSON array: {agendas.get(timeframe, agendas['weekly'])}"""
    try:
        raw = generate_ai_response(prompt, temperature=0.5, force_openai=True).replace('```json','').replace('```','').strip()
        return {"status": "success", "slides": json.loads(raw), "theme": theme_name}
    except Exception as e:
        print(f"Deck Parse Error: {e}", flush=True)
        return {"status": "error", "message": f"Failed to orchestrate {timeframe} slides."}

@app.post("/generate_ppt")
async def generate_ppt(payload: dict, creds: dict = Depends(get_jira_creds)):
    slides_data = payload.get("slides", [])
    theme_name  = payload.get("theme", "sprint")
    print(f"\U0001f3a8 Generating PPTX with theme: {theme_name} ({len(slides_data)} slides)", flush=True)
    ppt_buffer  = generate_native_editable_pptx(slides_data, theme_name)
    return StreamingResponse(ppt_buffer, headers={'Content-Disposition': f'attachment; filename="{payload.get("project","Project")}_Premium_Deck.pptx"'}, media_type="application/vnd.openxmlformats-officedocument.presentationml.presentation")

@app.get("/roadmap/{project_key}")
def get_roadmap(project_key: str, sprint_id: str = None, target_months: float = None, creds: dict = Depends(get_jira_creds)):
    """
    Strategic Roadmap — powered by Jira velocity + backlog.
    1. Fetches last N completed sprints + velocity
    2. Fetches backlog (not done, not in active sprint)
    3. Groups backlog into sprint-capacity buckets
    4. Runs strategic assembler (health score, investment buckets, delta, slippage)
    """
    sp_field = get_story_point_field(creds)

    # ── Step 1: Discover all sprints in this project ──
    sprint_discovery_res = jira_request("POST", "search/jql", creds, {
        "jql": f'project="{project_key}" AND sprint is not EMPTY ORDER BY updated DESC',
        "maxResults": 100,
        "fields": ["customfield_10020"]
    })
    all_sprints = {}
    if sprint_discovery_res and sprint_discovery_res.status_code == 200:
        for issue in sprint_discovery_res.json().get('issues', []):
            for s in (issue.get('fields') or {}).get('customfield_10020') or []:
                if s.get('id') not in all_sprints:
                    all_sprints[s['id']] = {
                        "id": s['id'], "name": s['name'], "state": s['state'],
                        "startDate": s.get('startDate', ''), "endDate": s.get('endDate', '')
                    }

    closed_sprints = sorted(
        [s for s in all_sprints.values() if s['state'] == 'closed'],
        key=lambda x: x.get('endDate', '') or str(x.get('id', 0)),
        reverse=True
    )[:8]
    closed_sprints.reverse()

    active_sprints = [s for s in all_sprints.values() if s['state'] == 'active']

    # ── Step 2: Calculate velocity from closed sprints ──
    safe_fields = ["summary", "status", "priority", "issuetype", "assignee",
                   "customfield_10016", "customfield_10026", "customfield_10028",
                   "customfield_10004", sp_field]

    sprint_history = []
    for sprint in closed_sprints:
        res = jira_request("POST", "search/jql", creds, {
            "jql": f'project="{project_key}" AND sprint={sprint["id"]}',
            "maxResults": 100,
            "fields": safe_fields
        })
        if not res or res.status_code != 200:
            continue

        issues = res.json().get('issues', [])
        completed_pts = 0.0
        total_pts = 0.0
        completed_count = 0
        total_count = len(issues)
        person_stats = {}

        for iss in issues:
            f = iss.get('fields') or {}
            pts = extract_story_points(f, sp_field)
            total_pts += pts
            status_cat = (f.get('status') or {}).get('statusCategory', {}).get('key', '')
            assignee_name = (f.get('assignee') or {}).get('displayName', 'Unassigned')

            if status_cat == 'done':
                completed_pts += pts
                completed_count += 1

            if assignee_name not in person_stats:
                person_stats[assignee_name] = {"completed": 0, "total": 0}
            person_stats[assignee_name]["total"] += pts
            if status_cat == 'done':
                person_stats[assignee_name]["completed"] += pts

        sprint_history.append({
            "name": sprint['name'],
            "total_issues": total_count,
            "completed_issues": completed_count,
            "total_points": total_pts,
            "velocity": completed_pts,
            "completed_points": completed_pts,
            "completion_rate": round((completed_count / max(total_count, 1)) * 100, 1),
            "person_stats": person_stats,
        })

    # Velocity metrics
    velocities = [s['velocity'] for s in sprint_history if s['velocity'] > 0]
    has_history = len(velocities) > 0

    if has_history:
        avg_velocity = round(sum(velocities) / len(velocities), 1)
        min_velocity = round(min(velocities), 1)
        max_velocity = round(max(velocities), 1)
    else:
        avg_velocity = 0
        min_velocity = 0
        max_velocity = 0

    trend = "no_data"
    if len(velocities) >= 4:
        early = sum(velocities[:2]) / 2
        recent = sum(velocities[-2:]) / 2
        if recent > early * 1.15: trend = "improving"
        elif recent < early * 0.85: trend = "declining"
        else: trend = "stable"
    elif has_history:
        trend = "stable"

    velocity_data = {
        "avg_velocity": avg_velocity,
        "min": min_velocity,
        "max": max_velocity,
        "trend": trend,
        "sprints_analyzed": len(velocities),
        "has_history": has_history,
    }

    # ── Step 3: Fetch backlog ──
    active_sprint_ids = set(str(s['id']) for s in active_sprints)
    backlog_jql = f'project="{project_key}" AND statusCategory != Done ORDER BY rank ASC, priority DESC'
    backlog_res = jira_request("POST", "search/jql", creds, {
        "jql": backlog_jql,
        "maxResults": 60,
        "fields": safe_fields + ["customfield_10020"]
    })

    backlog_items = []
    if backlog_res and backlog_res.status_code == 200:
        for iss in backlog_res.json().get('issues', []):
            f = iss.get('fields') or {}
            issue_sprints = f.get('customfield_10020') or []
            in_active = any(
                isinstance(sp, dict) and (str(sp.get('id', '')) in active_sprint_ids or sp.get('state') == 'active')
                for sp in issue_sprints
            )
            if in_active:
                continue
            pts = extract_story_points(f, sp_field)
            backlog_items.append({
                "key": iss.get('key'),
                "summary": f.get('summary', 'Untitled'),
                "points": pts,
                "type": (f.get('issuetype') or {}).get('name', 'Task'),
                "priority": (f.get('priority') or {}).get('name', 'Medium'),
                "status": (f.get('status') or {}).get('name', 'To Do'),
            })

    # If no history, estimate velocity from backlog
    if not has_history and backlog_items:
        total_backlog_pts = sum(item.get('points', 0) or 3 for item in backlog_items)
        avg_velocity = round(max(total_backlog_pts / 3, 10), 1)
        velocity_data["avg_velocity"] = avg_velocity
        velocity_data["estimated"] = True

    # ── Step 4: Generate sprint buckets by velocity capacity ──
    sprint_buckets = []
    if avg_velocity > 0 and backlog_items:
        current_bucket = {
            "sprint_label": "Sprint N+1",
            "target_capacity": avg_velocity,
            "allocated_points": 0,
            "items": []
        }
        bucket_idx = 1
        for item in backlog_items:
            pts = item.get("points", 0) or 3
            if current_bucket["allocated_points"] + pts > avg_velocity and current_bucket["items"]:
                sprint_buckets.append(current_bucket)
                bucket_idx += 1
                current_bucket = {
                    "sprint_label": f"Sprint N+{bucket_idx}",
                    "target_capacity": avg_velocity,
                    "allocated_points": 0,
                    "items": []
                }
            current_bucket["items"].append(item)
            current_bucket["allocated_points"] += pts
        if current_bucket["items"]:
            sprint_buckets.append(current_bucket)

    # ── Step 5: Assemble full strategic roadmap ──
    result = assemble_strategic_roadmap(
        project_key=project_key,
        sprint_history=sprint_history,
        velocity=velocity_data,
        backlog_items=backlog_items,
        sprint_buckets=sprint_buckets,
        target_months=target_months,
    )

    return result

@app.get("/capacity_check/{project_key}")
def capacity_check(project_key: str, sprint_id: str = None, creds: dict = Depends(get_jira_creds)):
    """
    Capacity Intelligence for Command Center.
    Compares current sprint committed points vs historical velocity.
    Returns warnings and per-person analysis.
    """
    sp_field = get_story_point_field(creds)

    # ── 1. Discover sprints ──
    sprint_disc = jira_request("POST", "search/jql", creds, {
        "jql": f'project="{project_key}" AND sprint is not EMPTY ORDER BY updated DESC',
        "maxResults": 100,
        "fields": ["customfield_10020"]
    })
    all_sprints = {}
    if sprint_disc and sprint_disc.status_code == 200:
        for issue in sprint_disc.json().get('issues', []):
            for s in (issue.get('fields') or {}).get('customfield_10020') or []:
                all_sprints[s['id']] = {"id": s['id'], "name": s['name'], "state": s['state']}

    closed_sprints = sorted(
        [s for s in all_sprints.values() if s['state'] == 'closed'],
        key=lambda x: x['id'], reverse=True
    )[:6]

    # ── 2. Calculate velocity from closed sprints ──
    safe_fields = ["summary", "status", "assignee", "priority",
                   "customfield_10016", "customfield_10026", "customfield_10028",
                   "customfield_10004", sp_field]

    sprint_velocities = []
    person_history = {}  # name -> [pts_per_sprint]

    for sprint in closed_sprints:
        res = jira_request("POST", "search/jql", creds, {
            "jql": f'project="{project_key}" AND sprint={sprint["id"]}',
            "maxResults": 100,
            "fields": safe_fields
        })
        if not res or res.status_code != 200:
            continue

        sprint_total = 0.0
        person_sprint = {}

        for iss in res.json().get('issues', []):
            f = iss.get('fields') or {}
            pts = extract_story_points(f, sp_field)
            status_cat = (f.get('status') or {}).get('statusCategory', {}).get('key', '')
            if status_cat == 'done':
                sprint_total += pts
                name = (f.get('assignee') or {}).get('displayName', 'Unassigned')
                person_sprint[name] = person_sprint.get(name, 0) + pts

        sprint_velocities.append(sprint_total)
        for name, pts in person_sprint.items():
            if name not in person_history:
                person_history[name] = []
            person_history[name].append(pts)

    avg_velocity = round(sum(sprint_velocities) / max(len(sprint_velocities), 1), 1) if sprint_velocities else 0
    person_avg = {name: round(sum(hist) / len(hist), 1) for name, hist in person_history.items()}
    has_history = len(sprint_velocities) > 0

    # ── 3. Get current sprint committed points ──
    jql = f'project="{project_key}" AND sprint={sprint_id}' if sprint_id and sprint_id != "active" else f'project="{project_key}" AND sprint in openSprints()'
    current_res = jira_request("POST", "search/jql", creds, {
        "jql": jql,
        "maxResults": 100,
        "fields": safe_fields
    })

    total_committed = 0.0
    total_completed = 0.0
    total_remaining = 0.0
    person_committed = {}
    person_completed = {}
    person_issues = {}

    if current_res and current_res.status_code == 200:
        for iss in current_res.json().get('issues', []):
            f = iss.get('fields') or {}
            pts = extract_story_points(f, sp_field)
            name = (f.get('assignee') or {}).get('displayName', 'Unassigned')
            status_cat = (f.get('status') or {}).get('statusCategory', {}).get('key', '')
            status_name = (f.get('status') or {}).get('name', 'To Do')
            priority_name = (f.get('priority') or {}).get('name', 'Medium')

            total_committed += pts
            person_committed[name] = person_committed.get(name, 0) + pts

            if name not in person_issues:
                person_issues[name] = []
            person_issues[name].append({
                "key": iss.get('key'),
                "summary": f.get('summary', ''),
                "points": pts,
                "status": status_name,
                "priority": priority_name
            })

            if status_cat == 'done':
                total_completed += pts
                person_completed[name] = person_completed.get(name, 0) + pts
            else:
                total_remaining += pts

    # ── 4. Calculate capacity utilization ──
    # When no history exists, use committed points as the baseline (first sprint mode)
    # This treats the current sprint commitment as the team's target velocity
    if has_history and avg_velocity > 0:
        utilization_pct = round((total_committed / avg_velocity) * 100, 1)
    elif total_committed > 0:
        # First sprint: show progress as utilization (completed/committed)
        utilization_pct = round((total_completed / total_committed) * 100, 1)
    else:
        utilization_pct = 0

    # Determine severity
    if not has_history and total_committed > 0:
        # First sprint mode — no history to compare against
        progress_pct = round((total_completed / total_committed) * 100, 1)
        severity = "first_sprint"
        warning = f"First sprint detected — no historical velocity to compare. The team has committed {total_committed} pts with {total_completed} completed ({progress_pct}% done). Complete this sprint to establish your velocity baseline."
    elif utilization_pct > 120:
        severity = "critical"
        warning = f"Sprint is significantly over-capacity at {utilization_pct}%! The team has committed {total_committed} points against an average velocity of {avg_velocity}. Consider removing {round(total_committed - avg_velocity, 1)} points of work."
    elif utilization_pct > 100:
        severity = "warning"
        warning = f"Sprint is over-capacity at {utilization_pct}%. The team has committed {total_committed} points but historically delivers ~{avg_velocity}. There's a risk of carryover."
    elif utilization_pct > 85:
        severity = "caution"
        warning = f"Sprint is at {utilization_pct}% capacity ({total_committed}/{avg_velocity} pts). This is ambitious but achievable if the team maintains focus."
    elif utilization_pct > 0:
        severity = "healthy"
        warning = f"Sprint is at {utilization_pct}% capacity ({total_committed}/{avg_velocity} pts). The team has room for additional work."
    else:
        severity = "empty"
        warning = "No active sprint or no committed work found."

    # Per-person breakdown
    # When no history: calculate each person's share of the sprint
    num_members = len([n for n in person_committed.keys() if n != 'Unassigned']) or 1
    fair_share = round(total_committed / num_members, 1) if not has_history and total_committed > 0 else 0

    person_capacity = {}
    for name in set(list(person_committed.keys()) + list(person_avg.keys())):
        committed = person_committed.get(name, 0)
        completed = person_completed.get(name, 0)
        hist_avg = person_avg.get(name, 0)

        if has_history and hist_avg > 0:
            util = round((committed / hist_avg) * 100, 1)
        elif committed > 0:
            # First sprint: show completion % as utilization
            util = round((completed / committed) * 100, 1)
        else:
            util = 0

        if has_history:
            person_status = "healthy"
            if util > 120:
                person_status = "overloaded"
            elif util > 100:
                person_status = "at_risk"
            elif util < 50 and hist_avg > 0:
                person_status = "underutilized"
        else:
            # First sprint: flag based on workload distribution
            person_status = "first_sprint"
            if name != 'Unassigned' and fair_share > 0:
                load_ratio = committed / fair_share
                if load_ratio > 1.5:
                    person_status = "heavy_load"
                elif load_ratio < 0.5 and committed > 0:
                    person_status = "light_load"

        person_capacity[name] = {
            "committed": committed,
            "completed": completed,
            "remaining": committed - completed,
            "historical_avg": hist_avg,
            "fair_share": fair_share if not has_history else 0,
            "utilization_pct": util,
            "status": person_status,
            "issues": person_issues.get(name, [])
        }

    # ── 5. Recommendations ──
    recommendations = []

    if not has_history and total_committed > 0:
        # First sprint recommendations
        progress_pct = round((total_completed / total_committed) * 100, 1)
        recommendations.append(f"Sprint progress: {progress_pct}% complete ({total_completed}/{total_committed} pts).")

        heavy = [n for n, c in person_capacity.items() if c["status"] == "heavy_load" and n != "Unassigned"]
        light = [n for n, c in person_capacity.items() if c["status"] == "light_load" and n != "Unassigned"]
        if heavy:
            recommendations.append(f"Uneven distribution: {', '.join(heavy)} {'have' if len(heavy) > 1 else 'has'} significantly more work than the team average ({fair_share} pts).")
        if light:
            recommendations.append(f"{', '.join(light)} {'have' if len(light) > 1 else 'has'} lighter workload — consider rebalancing.")
        recommendations.append("Complete this sprint to establish historical velocity for future capacity planning.")
    else:
        overloaded = [n for n, c in person_capacity.items() if c["status"] == "overloaded" and n != "Unassigned"]
        underutilized = [n for n, c in person_capacity.items() if c["status"] == "underutilized" and n != "Unassigned"]

        if overloaded:
            recommendations.append(f"{'These members are' if len(overloaded) > 1 else 'This member is'} over-capacity: {', '.join(overloaded)}. Consider redistributing work.")
        if underutilized:
            recommendations.append(f"{'These members have' if len(underutilized) > 1 else 'This member has'} available bandwidth: {', '.join(underutilized)}.")
        if severity in ["critical", "warning"]:
            excess = round(total_committed - avg_velocity, 1)
            recommendations.append(f"Remove approximately {excess} story points from this sprint to match team velocity.")
        if total_remaining > 0 and total_completed > 0:
            progress_pct = round((total_completed / total_committed) * 100, 1)
            recommendations.append(f"Sprint progress: {progress_pct}% complete ({total_completed}/{total_committed} pts).")

    return {
        "severity": severity,
        "warning": warning,
        "utilization_pct": utilization_pct,
        "total_committed": total_committed,
        "total_completed": total_completed,
        "total_remaining": total_remaining,
        "avg_velocity": avg_velocity,
        "velocity_history": sprint_velocities,
        "person_capacity": person_capacity,
        "recommendations": recommendations,
        "sprints_analyzed": len(sprint_velocities),
        "has_history": has_history
    }


@app.post("/timeline/generate_story")
async def generate_timeline_story(payload: dict, creds: dict = Depends(get_jira_creds)):
    sp_field = get_story_point_field(creds); project_key = payload.get('project')
    roster, assignable_map = build_team_roster(project_key, creds, sp_field)
    res = jira_request("POST", "search/jql", creds, {"jql": f'project="{project_key}" AND sprint in openSprints()', "fields": ["summary", "assignee"]})
    board_context = []
    if res is not None and res.status_code == 200:
        for i in res.json().get('issues', []):
            f = i.get('fields') or {}; name = (f.get('assignee') or {}).get('displayName')
            board_context.append(f"Task: {f.get('summary')} | Assignee: {name or 'Unassigned'}")
    prompt_text = f"Product Owner. User Request: '{payload.get('prompt')}'. Current Sprint Context: {' '.join(board_context[:20])}. Valid Team Roster (MUST pick EXACT NAME from keys): {json.dumps(roster)}. Generate a detailed user story. Return JSON: {{\"title\": \"...\", \"description\": \"...\", \"acceptance_criteria\": [\"...\"], \"points\": 5, \"assignee\": \"Exact Name\", \"tech_stack_inferred\": \"...\"}}"
    try:
        raw_response = generate_ai_response(prompt_text, temperature=0.5, image_data=payload.get("image_data"))
        if not raw_response: return {"status": "error", "message": "AI model failed to generate response."}
        return {"status": "success", "story": json.loads(raw_response.replace('```json','').replace('```','').strip())}
    except Exception as e: return {"status": "error", "message": str(e)}

@app.post("/timeline/generate_epic")
async def generate_epic(payload: dict, creds: dict = Depends(get_jira_creds)):
    project_key = payload.get('project')
    res = jira_request("POST", "search/jql", creds, {"jql": f'project="{project_key}"', "maxResults": 10, "fields": ["summary"]})
    board_context = []
    if res is not None and res.status_code == 200:
        for i in res.json().get('issues', []): board_context.append((i.get('fields') or {}).get('summary', ''))
    prompt_text = f"Chief Product Officer. User Input: '{payload.get('prompt')}'. Project Context: {json.dumps(board_context)}. Transform into implementation-ready Agile Epic. Return STRICT JSON: {{\"title\": \"Epic Name\", \"motivation\": \"Why are we building this?\", \"description\": \"Detailed scope.\", \"acceptance_criteria\": [\"AC1\", \"AC2\"]}}"
    try:
        raw_response = generate_ai_response(prompt_text, temperature=0.6, force_openai=True)
        if not raw_response: return {"status": "error", "message": "AI model failed to generate epic."}
        return {"status": "success", "epic": json.loads(raw_response.replace('```json','').replace('```','').strip())}
    except Exception as e: return {"status": "error", "message": str(e)}

@app.post("/timeline/create_issue")
async def create_issue(payload: dict, creds: dict = Depends(get_jira_creds)):
    story = payload.get("story", {})
    project_key = payload.get("project")
    issue_type = payload.get("issue_type", "Story")
    sp_field = get_story_point_field(creds)
    roster, assignable_map = build_team_roster(project_key, creds, sp_field)
    target_assignee = story.get("assignee", "")
    assignee_id = assignable_map.get(target_assignee)
    if not assignee_id and target_assignee and target_assignee.lower() != "unassigned":
        assignee_id = get_jira_account_id(target_assignee, creds)
    base_url = ""
    if creds.get("auth_type") == "basic": base_url = f"https://{creds['domain']}"
    else:
        server_info = jira_request("GET", "serverInfo", creds)
        if server_info is not None and server_info.status_code == 200: base_url = server_info.json().get("baseUrl", "")
    desc_text = story.get("description", "AI Generated Item")
    if issue_type == "Epic" and story.get("motivation"):
        desc_text = f"Motivation:\n{story.get('motivation')}\n\nDescription:\n{desc_text}"
    issue_data = {"fields": {"project": {"key": project_key}, "summary": story.get("title", f"AI Generated {issue_type}"), "description": create_adf_doc(desc_text, story.get("acceptance_criteria")), "issuetype": {"name": issue_type}}}
    res = jira_request("POST", "issue", creds, issue_data)
    if res is None or res.status_code != 201:
        if issue_type == "Epic":
            issue_data["fields"]["issuetype"]["name"] = "Story"
            res = jira_request("POST", "issue", creds, issue_data)
        if res is None or res.status_code != 201:
            issue_data["fields"]["issuetype"]["name"] = "Task"
            res_fallback = jira_request("POST", "issue", creds, issue_data)
            if res_fallback is not None: res = res_fallback
    if res is not None and res.status_code == 201:
        new_key = res.json().get("key")
        if assignee_id: jira_request("PUT", f"issue/{new_key}", creds, {"fields": {"assignee": {"accountId": assignee_id}}})
        points = safe_float(story.get("points", 0))
        if points > 0 and issue_type != "Epic":
            pts_res = jira_request("PUT", f"issue/{new_key}", creds, {"fields": {sp_field: points}})
            if pts_res is not None and pts_res.status_code not in [200, 204]: print(f"Warning: Could not set Story Points on {new_key}.", flush=True)
        if issue_type != "Epic":
            jira_request("POST", f"issue/{new_key}/comment", creds, {"body": {"type": "doc", "version": 1, "content": [{"type": "paragraph", "content": [{"type": "text", "text": f"IG Agile AI Insights:\n- Estimation: {story.get('points', 0)} pts.\n- Reasoning: {story.get('tech_stack_inferred', '')}"}]}]}})
        issue_url = f"{base_url}/browse/{new_key}" if base_url else f"https://id.atlassian.com/browse/{new_key}"
        return {"status": "success", "key": new_key, "url": issue_url}
    error_message = extract_jira_error(res)
    return {"status": "error", "message": error_message}

@app.get("/reports/{project_key}/{timeframe}")
def get_report(project_key: str, timeframe: str, creds: dict = Depends(get_jira_creds)):
    sp_field = get_story_point_field(creds)
    days = 7 if timeframe == "weekly" else (30 if timeframe == "monthly" else 90)
    dt = (datetime.now() - timedelta(days=days)).strftime("%Y-%m-%d")
    safe_fields = ["summary", "status", "assignee", "priority", sp_field]
    res = jira_request("POST", "search/jql", creds, {"jql": f'project="{project_key}" AND updated >= "{dt}" ORDER BY updated DESC', "maxResults": 40, "fields": safe_fields})
    done_count = 0; done_pts = 0.0; context_data = []
    for i in res.json().get('issues', []) if res is not None and res.status_code == 200 else []:
        f = i.get('fields') or {}
        pts = extract_story_points(f, sp_field)
        if (f.get('status') or {}).get('statusCategory', {}).get('key') == 'done':
            done_count += 1; done_pts += pts
        status_name = (f.get('status') or {}).get('name') or "Unknown"
        assignee = (f.get('assignee') or {}).get('displayName') or "Unassigned"
        context_data.append({"key": i.get('key'), "summary": f.get('summary', ''), "status": status_name, "assignee": assignee, "points": pts})
    try:
        raw = generate_ai_response(f"Elite Agile Analyst. DATA: {json.dumps(context_data)}. Return JSON: {{\"ai_verdict\": \"...\", \"sprint_vibe\": \"...\", \"key_accomplishments\": [{{\"title\": \"...\", \"impact\": \"...\"}}], \"hidden_friction\": \"...\", \"top_contributor\": \"Name - Reason\"}}", temperature=0.4).replace('```json','').replace('```','').strip()
        ai_dossier = json.loads(raw)
    except Exception as e:
        ai_dossier = {"ai_verdict": "Error analyzing data.", "sprint_vibe": "Error", "key_accomplishments": [], "hidden_friction": "", "top_contributor": ""}
    return {"completed_count": done_count, "completed_points": done_pts, "total_active_in_period": len(context_data), "dossier": ai_dossier}

@app.get("/retro/{project_key}")
def get_retro(project_key: str, sprint_id: str, creds: dict = Depends(get_jira_creds)):
    res = jira_request("GET", f"project/{project_key}/properties/ig_agile_retro", creds)
    db_data = res.json().get('value', {}) if res is not None and res.status_code == 200 else {}
    if str(sprint_id) not in db_data: db_data[str(sprint_id)] = {"well": [], "improve": [], "kudos": [], "actions": []}
    return db_data[str(sprint_id)]

@app.post("/retro/update")
def update_retro(payload: dict, creds: dict = Depends(get_jira_creds)):
    project_key = payload.get("project").upper(); sid = str(payload.get("sprint")); res = jira_request("GET", f"project/{project_key}/properties/ig_agile_retro", creds)
    db_data = res.json().get('value', {}) if res is not None and res.status_code == 200 else {}; db_data[sid] = payload.get("board")
    jira_request("PUT", f"project/{project_key}/properties/ig_agile_retro", creds, db_data)
    return {"status": "saved"}

@app.post("/retro/chat")
def retro_chat(payload: dict):
    board = payload.get('board', {})
    question = payload.get('question', '')
    well = [item.get('text') for item in board.get('well', [])]
    improve = [item.get('text') for item in board.get('improve', [])]
    kudos = [item.get('text') for item in board.get('kudos', [])]
    prompt = f"You are an Expert Agile Coach assisting a Scrum Master. Board data: WENT WELL: {well} | NEEDS IMPROVEMENT: {improve} | KUDOS: {kudos}. Question: '{question}'. Provide a concise, highly actionable response using ONLY the provided data. Use markdown formatting. Do not output JSON."
    try: return {"reply": generate_ai_response(prompt, temperature=0.5, force_openai=True, json_mode=False)}
    except Exception as e: print(f"Retro Chat Error: {e}", flush=True); return {"reply": "I encountered an error connecting to the AI network."}

@app.post("/guest/retro/generate_link")
def generate_retro_link(payload: dict, creds: dict = Depends(get_jira_creds), db: Session = Depends(get_db)):
    project_key = payload.get("project")
    sprint_id = payload.get("sprint")
    license_key = payload.get("license_key")
    if not license_key: raise HTTPException(status_code=400, detail="Missing license key")
    token = str(uuid.uuid4())
    db.add(GuestLink(token=token, project_key=project_key, sprint_id=str(sprint_id), license_key=license_key))
    db.commit()
    return {"token": token}

@app.get("/guest/retro/{token}")
def get_guest_retro(token: str, db: Session = Depends(get_db)):
    link = db.query(GuestLink).filter(GuestLink.token == token).first()
    if not link: raise HTTPException(status_code=404, detail="Invalid or expired link")
    user = get_valid_oauth_session(db=db, license_key=link.license_key)
    if not user: raise HTTPException(status_code=401, detail="Host session expired. Please ask the Scrum Master to generate a new link.")
    creds = {"auth_type": "oauth", "cloud_id": user.cloud_id, "access_token": user.access_token}
    res = jira_request("GET", f"project/{link.project_key}/properties/ig_agile_retro", creds)
    db_data = res.json().get('value', {}) if res and res.status_code == 200 else {}
    board = db_data.get(link.sprint_id, {"well": [], "improve": [], "kudos": [], "actions": []})
    return {"project": link.project_key, "sprint": link.sprint_id, "board": board}

@app.post("/guest/retro/{token}/add")
def add_guest_retro(token: str, payload: dict, db: Session = Depends(get_db)):
    link = db.query(GuestLink).filter(GuestLink.token == token).first()
    if not link: raise HTTPException(status_code=404, detail="Invalid link")
    user = get_valid_oauth_session(db=db, license_key=link.license_key)
    if not user: raise HTTPException(status_code=401, detail="Host session expired")
    creds = {"auth_type": "oauth", "cloud_id": user.cloud_id, "access_token": user.access_token}
    res = jira_request("GET", f"project/{link.project_key}/properties/ig_agile_retro", creds)
    db_data = res.json().get('value', {}) if res and res.status_code == 200 else {}
    if link.sprint_id not in db_data:
        db_data[link.sprint_id] = {"well": [], "improve": [], "kudos": [], "actions": []}
    col = payload.get("column")
    text_val = payload.get("text")
    if col in ['well', 'improve', 'kudos'] and text_val:
        db_data[link.sprint_id][col].append({"id": int(time.time()*1000), "text": text_val})
        jira_request("PUT", f"project/{link.project_key}/properties/ig_agile_retro", creds, db_data)
    return {"status": "success", "board": db_data[link.sprint_id]}

# ================= TEAM & ROLES MANAGEMENT =================

@app.get("/team/{project_key}")
def get_team(project_key: str, creds: dict = Depends(get_jira_creds)):
    """Get saved team roles from Jira project properties."""
    res = jira_request("GET", f"project/{project_key}/properties/ig_agile_team", creds)
    team_data = res.json().get('value', {}) if res is not None and res.status_code == 200 else {}
    team = team_data.get('members', [])
    
    # Also get assignable users for quick-add
    jira_users = []
    try:
        user_res = jira_request("GET", f"user/assignable/search?project={project_key}&maxResults=50", creds)
        if user_res is not None and user_res.status_code == 200:
            for u in user_res.json():
                if u.get('displayName') and u.get('accountType') == 'atlassian':
                    jira_users.append({
                        "displayName": u.get('displayName', ''),
                        "accountId": u.get('accountId', ''),
                        "email": u.get('emailAddress', ''),
                        "avatar": u.get('avatarUrls', {}).get('48x48', '')
                    })
    except Exception as e:
        print(f"Error fetching Jira users: {e}", flush=True)
    
    return {"team": team, "jira_users": jira_users}

@app.get("/team/{project_key}/fetch_jira")
def fetch_jira_team(project_key: str, creds: dict = Depends(get_jira_creds)):
    """Fetch assignable users from Jira for team building."""
    jira_users = []
    try:
        user_res = jira_request("GET", f"user/assignable/search?project={project_key}&maxResults=50", creds)
        if user_res is not None and user_res.status_code == 200:
            for u in user_res.json():
                if u.get('displayName') and u.get('accountType') == 'atlassian':
                    jira_users.append({
                        "displayName": u.get('displayName', ''),
                        "accountId": u.get('accountId', ''),
                        "email": u.get('emailAddress', ''),
                        "avatar": u.get('avatarUrls', {}).get('48x48', '')
                    })
    except Exception as e:
        print(f"Error fetching Jira users: {e}", flush=True)
    
    return {"jira_users": jira_users}

@app.post("/team/{project_key}/save")
def save_team(project_key: str, payload: dict, creds: dict = Depends(get_jira_creds)):
    """Save team roles to Jira project properties for persistence."""
    team_members = payload.get("team", [])
    team_data = {"members": team_members, "updated_at": datetime.utcnow().isoformat()}
    
    jira_request("PUT", f"project/{project_key.upper()}/properties/ig_agile_team", creds, team_data)
    return {"status": "saved", "count": len(team_members)}

@app.post("/feature_roadmap")
def generate_feature_roadmap(req: FeatureRoadmapRequest, creds: dict = Depends(get_jira_creds)):

    # ── Normalize target duration to working days ──
    unit = req.target_duration_unit.lower()
    val = req.target_duration_value
    if unit == "days":
        target_working_days = val
    elif unit == "months":
        target_working_days = int(val * 22)
    elif unit == "years":
        target_working_days = int(val * 260)
    else:
        target_working_days = int(val * 22)

    target_months = round(target_working_days / 22, 1)
    target_sprints = math.ceil(target_working_days / 10)

    num_features = len(req.features)
    start_date = req.start_date or datetime.now().strftime('%Y-%m-%d')

    print(f"\n{'='*60}", flush=True)
    print(f"🗺️  Feature Roadmap: {num_features} features | Target: {target_months}m", flush=True)
    print(f"   Tech: {req.tech_stack} | Start: {start_date}", flush=True)
    print(f"{'='*60}\n", flush=True)

    # ── CHOOSE MODEL: Premium for roadmap accuracy ──
    # o3-mini for deep reasoning, gpt-4o as fallback
    ROADMAP_MODEL = os.getenv("ROADMAP_AI_MODEL", "gemini-3.1-pro")
    ROADMAP_TIMEOUT = 300  # seconds per AI call
    FEATURE_BATCH_SIZE = 20

    # ── CHUNKING STRATEGY ──
    # Small (≤20): single call — everything in one shot
    # Medium (21-40): 2-phase — structure first, then epics
    # Large (41+): 3-phase — structure, epics batch 1, epics batch 2+
    SINGLE_CALL_LIMIT = 20
    EPIC_BATCH_SIZE = 15

    features_text = "\n".join([f"{i+1}. {f}" for i, f in enumerate(req.features)])

    if num_features <= SINGLE_CALL_LIMIT:
        # ═══ SINGLE CALL MODE (≤20 features) ═══
        print(f"📦 Single-call mode ({num_features} features)", flush=True)
        return _roadmap_single_call(
            req, features_text, num_features, start_date,
            target_working_days, target_months, target_sprints,
            ROADMAP_MODEL, ROADMAP_TIMEOUT
        )
    else:
         print(f"🔀 Chunked mode ({num_features} features → {math.ceil(num_features/FEATURE_BATCH_SIZE)} structure batches + {math.ceil(num_features/EPIC_BATCH_SIZE)} epic batches)", flush=True)
         return _roadmap_chunked(
             req, features_text, num_features, start_date,
             target_working_days, target_months, target_sprints,
             ROADMAP_MODEL, ROADMAP_TIMEOUT, EPIC_BATCH_SIZE, FEATURE_BATCH_SIZE
         )


def _roadmap_single_call(req, features_text, num_features, start_date,
                          target_working_days, target_months, target_sprints,
                          model, timeout):
    """Original single-call approach for ≤20 features."""
    prompt = _build_full_roadmap_prompt(
        req.tech_stack, features_text, num_features, start_date,
        target_working_days, target_months, target_sprints
    )

    try:
        ai_result = generate_ai_response(prompt, temperature=0.2, timeout=timeout, model=model)
        if not ai_result:
            raise ValueError("AI returned empty response")

        raw = ai_result.replace('```json', '').replace('```', '').strip()
        parsed = json.loads(raw)
        return _post_process_roadmap(parsed, req, start_date, target_working_days, target_months, target_sprints, num_features)

    except json.JSONDecodeError as e:
        print(f"Feature Roadmap JSON Error: {e}", flush=True)
        return {"error": f"AI response was not valid JSON. Please try again.", "raw_snippet": str(e)}
    except Exception as e:
        print(f"Feature Roadmap Error: {e}", flush=True)
        return {"error": str(e)}

def _roadmap_chunked(req, features_text, num_features, start_date,
                      target_working_days, target_months, target_sprints,
                      model, timeout, epic_batch_size, feature_batch_size=20):
    """
    3-Phase chunked approach for >20 features:
      Phase 1A: Feature analysis in batches of ~20
      Phase 1B: Project structure (team, timeline, gantt, sprint_map) from aggregated features
      Phase 2:  Epics & stories in batches of ~15
    """

    features_list = req.features

    # ═══════════════════════════════════════════════════════════
    # PHASE 1A: Feature Analysis in batches
    # ═══════════════════════════════════════════════════════════
    all_feature_analysis = []
    total_fa_batches = math.ceil(num_features / feature_batch_size)

    for batch_idx in range(total_fa_batches):
        batch_start = batch_idx * feature_batch_size
        batch_end = min(batch_start + feature_batch_size, num_features)
        batch_features = features_list[batch_start:batch_end]
        batch_count = len(batch_features)

        batch_text = "\n".join([f"{batch_start + i + 1}. {f}" for i, f in enumerate(batch_features)])

        print(f"  ⚡ Phase 1A Batch {batch_idx+1}/{total_fa_batches}: Analyzing features {batch_start+1}-{batch_end}...", flush=True)

        fa_prompt = f"""You are a world-class Technical Project Manager with 20+ years experience.

CONTEXT:
- Tech Stack: {req.tech_stack}
- These are features {batch_start+1} to {batch_end} out of {num_features} total features
- Project Start Date: {start_date}

FEATURES TO ANALYZE:
{batch_text}

SIZING RULES: T-Shirt → Fibonacci ONLY: XXS=1, SMALL=2, MEDIUM=3, LARGE=5, XL=8, XXL=13, XXXL=21
Feature types: "Application (UI)", "Systems Integration", "Reporting", "Process Automation", "Data & Backend"

For EACH feature, provide detailed analysis. Return ONLY valid JSON (no markdown):
{{
    "feature_analysis": [
        {{
            "id": {batch_start + 1},
            "feature": "Exact feature name",
            "feature_type": "Application (UI)",
            "technical_scope": "Detailed technical scope for {req.tech_stack} (2-3 sentences)",
            "size": "XL",
            "story_points": 8,
            "days": 8,
            "roles_needed": "PP, QA, BA",
            "est_team": "XL – PP, QA, BA",
            "est_conservative": "L – PP, QA",
            "dependencies": "Feature #{batch_start + 1}" 
        }}
    ]
}}

RULES:
1. ALL {batch_count} features MUST appear in feature_analysis
2. IDs must be sequential starting from {batch_start + 1}
3. Be specific to {req.tech_stack} in technical_scope
4. Dependencies reference feature # numbers from the FULL list (1-{num_features})"""

        try:
            result = generate_ai_response(fa_prompt, temperature=0.2, timeout=timeout, model=model)
            if result:
                raw = result.replace('```json', '').replace('```', '').strip()
                parsed_batch = json.loads(raw)
                batch_fa = parsed_batch.get("feature_analysis", [])
                all_feature_analysis.extend(batch_fa)
                print(f"  ✅ Phase 1A Batch {batch_idx+1}: {len(batch_fa)} features analyzed", flush=True)
            else:
                print(f"  ⚠️ Phase 1A Batch {batch_idx+1}: Empty response, creating defaults", flush=True)
                for i, f in enumerate(batch_features):
                    all_feature_analysis.append({
                        "id": batch_start + i + 1, "feature": f,
                        "feature_type": "Application (UI)",
                        "technical_scope": f"Implementation of {f} using {req.tech_stack}",
                        "size": "MEDIUM", "story_points": 3, "days": 3,
                        "roles_needed": "PP, QA", "est_team": "M – PP, QA",
                        "est_conservative": "S – PP", "dependencies": "None"
                    })
        except Exception as e:
            print(f"  ⚠️ Phase 1A Batch {batch_idx+1} Error: {e}. Creating defaults.", flush=True)
            for i, f in enumerate(batch_features):
                all_feature_analysis.append({
                    "id": batch_start + i + 1, "feature": f,
                    "feature_type": "Application (UI)",
                    "technical_scope": f"Implementation of {f} using {req.tech_stack}",
                    "size": "MEDIUM", "story_points": 3, "days": 3,
                    "roles_needed": "PP, QA", "est_team": "M – PP, QA",
                    "est_conservative": "S – PP", "dependencies": "None"
                })

    print(f"\n  📊 Phase 1A complete: {len(all_feature_analysis)} features analyzed", flush=True)

    # ═══════════════════════════════════════════════════════════
    # PHASE 1B: Project Structure (using aggregated feature data)
    # ═══════════════════════════════════════════════════════════
    print(f"  ⚡ Phase 1B: Generating project structure...", flush=True)

    total_pts = sum(f.get("story_points", 3) for f in all_feature_analysis)
    total_seq_days = sum(f.get("days", 3) for f in all_feature_analysis)

    # Build compact feature summary for the structure prompt
    feature_summary = json.dumps([
        {"id": f.get("id"), "feature": f.get("feature", "")[:60], "pts": f.get("story_points", 3),
         "days": f.get("days", 3), "type": f.get("feature_type", ""), "deps": f.get("dependencies", "None")}
        for f in all_feature_analysis
    ])

    structure_prompt = f"""You are an expert Release Train Engineer and Project Planner.

CONTEXT:
- Tech Stack: {req.tech_stack}
- Total Features: {num_features}
- Total Story Points: {total_pts}
- Total Sequential Days: {total_seq_days}
- Client Target: {target_working_days} working days ({target_months} months)
- Start Date: {start_date}

FEATURE SUMMARY (already analyzed):
{feature_summary}

TASK: Generate the project structure around these features. Calculate scheduling, team, and timeline.

TARGET DURATION SCALING:
- Sequential effort: {total_seq_days} days
- Target: {target_working_days} days
- If sequential > target: need parallel streams (max 4)
- Each stream: separate dev, QA, BA roles + 10% coordination overhead per extra stream

Return ONLY valid JSON (no markdown):
{{
    "sizing_legend": [
        {{"size": "XXS", "days": 1, "story_points": 1, "sprints_equivalent": "0.1 sprint"}},
        {{"size": "SMALL", "days": 2, "story_points": 2, "sprints_equivalent": "0.2 sprint"}},
        {{"size": "MEDIUM", "days": 3, "story_points": 3, "sprints_equivalent": "0.3 sprint"}},
        {{"size": "LARGE", "days": 5, "story_points": 5, "sprints_equivalent": "0.5 sprint"}},
        {{"size": "XL", "days": 8, "story_points": 8, "sprints_equivalent": "0.8 sprint"}},
        {{"size": "XXL", "days": 13, "story_points": 13, "sprints_equivalent": "1.3 sprints"}},
        {{"size": "XXXL", "days": 21, "story_points": 21, "sprints_equivalent": "2.1 sprints"}}
    ],
    "scheduling": [
        {{"id": 1, "start_day": 1, "end_day": 8, "sprint_allocation": "SP1"}}
    ],
    "team_composition": [
        {{
            "role": "Program Manager (PgM)",
            "headcount": 1,
            "billable": true,
            "justification": "Overall delivery coordination",
            "ramp_up_notes": "From Day 1",
            "stream_allocation": "All streams"
        }}
    ],
    "total_team_size": 0,
    "parallel_stream_analysis": {{
        "single_stream_days": {total_seq_days},
        "single_stream_months": {round(total_seq_days / 22, 1)},
        "recommended_streams": 0,
        "actual_parallel_days": 0,
        "actual_parallel_months": 0,
        "target_days": {target_working_days},
        "target_months": {target_months},
        "fits_target": true,
        "coordination_overhead_pct": 0,
        "notes": "Analysis based on {num_features} features totaling {total_pts} story points"
    }},
    "timeline": {{
        "total_story_points": {total_pts},
        "team_velocity_per_sprint": 0,
        "sprint_duration_weeks": 2,
        "total_sprints": 0,
        "total_working_days": 0,
        "total_months": 0,
        "start_date": "{start_date}",
        "end_date": "YYYY-MM-DD",
        "assumptions": "Based on N parallel streams"
    }},
    "sprint_mapping": [
        {{
            "sprint": "SP1", "start_day": 1, "end_day": 10,
            "month": "M1", "calendar_month": "Mar 2026",
            "features_in_sprint": ["Feature #1", "Feature #2"],
            "points_in_sprint": 0
        }}
    ],
    "gantt_phases": [
        {{
            "phase": "Planning & Discovery",
            "assigned_roles": "PgM, Architect",
            "dependencies": "None",
            "start_day": 1, "end_day": 5,
            "start_week": 1, "end_week": 1,
            "duration_days": 5,
            "phase_type": "planning"
        }}
    ],
    "resource_loading": [
        {{"day": 1, "sprint": "SP1", "month": "M1", "active_features": 2, "team_members_needed": 4, "roles_active": "PP, QA, BA, PM"}}
    ],
    "uat_milestones": [
        {{
            "name": "UAT 1", "sprint": "SP4", "day": 40,
            "scope": "Core modules", "duration_days": 5,
            "exit_criteria": "All P1 defects resolved"
        }}
    ],
    "pilot_hypercare": {{
        "pilot": {{"start_day": 0, "end_day": 0, "duration_days": 15, "description": "Limited production deployment", "team_needed": "PM, 1 Dev, 1 QA, 1 BA"}},
        "hypercare": {{"start_day": 0, "end_day": 0, "duration_days": 15, "description": "Full production support", "team_needed": "PM, 2 Dev, 1 QA"}}
    }},
    "feature_type_summary": {{}}
}}

RULES:
1. "scheduling" array: one entry per feature with start_day, end_day, sprint_allocation
   - Respect dependencies (dependent features start AFTER their dependency ends)
   - Parallel features can overlap if they don't share roles
2. Sprint mapping: SP1=Day 1-10, SP2=Day 11-20, etc. Calendar months from {start_date}
3. UAT gates every 4-6 sprints + final UAT before pilot
4. Pilot 15 days after final UAT, Hypercare 15 days after pilot
5. resource_loading: entries for every 10th day
6. gantt_phases: Planning, Setup, Development sprints, QA, each UAT, Pilot, Hypercare
7. feature_type_summary: count and points per feature type"""

    try:
        structure_result = generate_ai_response(structure_prompt, temperature=0.2, timeout=timeout, model=model)
        if not structure_result:
            raise ValueError("Phase 1B AI returned empty response")

        structure_raw = structure_result.replace('```json', '').replace('```', '').strip()
        structure = json.loads(structure_raw)
        print(f"  ✅ Phase 1B complete: project structure generated", flush=True)

    except Exception as e:
        print(f"  ⚠️ Phase 1B Error: {e}. Building fallback structure.", flush=True)
        structure = _build_fallback_structure(
            all_feature_analysis, num_features, total_pts, total_seq_days,
            target_working_days, target_months, start_date
        )

    # ═══ MERGE Phase 1A + 1B ═══
    # Apply scheduling data back to feature_analysis
    scheduling = {s.get("id"): s for s in structure.get("scheduling", [])}
    for fa in all_feature_analysis:
        sched = scheduling.get(fa.get("id"), {})
        fa["start_day"] = sched.get("start_day", 1)
        fa["end_day"] = sched.get("end_day", fa.get("start_day", 1) + fa.get("days", 3) - 1)
        fa["sprint_allocation"] = sched.get("sprint_allocation", "SP1")

    parsed = structure
    parsed["feature_analysis"] = all_feature_analysis
    parsed["total_story_points"] = total_pts
    parsed["total_working_days_sequential"] = total_seq_days

    # ═══════════════════════════════════════════════════════════
    # PHASE 2: Epics & Stories (unchanged from before)
    # ═══════════════════════════════════════════════════════════
    all_epics = []
    total_epic_batches = math.ceil(len(all_feature_analysis) / epic_batch_size)

    for batch_idx in range(total_epic_batches):
        batch_start = batch_idx * epic_batch_size
        batch_end = min(batch_start + epic_batch_size, len(all_feature_analysis))
        batch_features = all_feature_analysis[batch_start:batch_end]

        print(f"  ⚡ Phase 2 Batch {batch_idx+1}/{total_epic_batches}: Epics for features {batch_start+1}-{batch_end}...", flush=True)

        batch_features_text = "\n".join([
            f"- {f.get('feature', '')} (ID:{f.get('id','')}, {f.get('story_points',3)} pts, {f.get('feature_type','Application (UI)')})"
            for f in batch_features
        ])

        epic_prompt = f"""You are an expert Agile Product Owner. Generate Jira-ready Epics with User Stories.

Tech Stack: {req.tech_stack}

FEATURES TO BREAK DOWN INTO EPICS:
{batch_features_text}

For EACH feature above, create ONE epic with 1-4 user stories.

Return ONLY valid JSON (no markdown):
{{
    "epics": [
        {{
            "epic_name": "Feature Name as Epic",
            "epic_description": "What this epic covers",
            "feature_type": "Application (UI)",
            "total_points": 8,
            "stories": [
                {{
                    "summary": "Implement [specific functionality]",
                    "description": "As a [user], I want [feature] so that [benefit].\\n\\nAcceptance Criteria:\\n- Criterion 1\\n- Criterion 2",
                    "story_points": 3,
                    "priority": "High"
                }}
            ]
        }}
    ]
}}

RULES:
1. One epic per feature — {len(batch_features)} epics total
2. Story points within an epic must sum to the feature's total points
3. Each story needs clear acceptance criteria
4. Priorities: High for core features, Medium for enhancements"""

        try:
            epic_result = generate_ai_response(epic_prompt, temperature=0.3, timeout=timeout, model=model)
            if epic_result:
                epic_raw = epic_result.replace('```json', '').replace('```', '').strip()
                epic_parsed = json.loads(epic_raw)
                batch_epics = epic_parsed.get("epics", [])
                all_epics.extend(batch_epics)
                print(f"  ✅ Batch {batch_idx+1}: {len(batch_epics)} epics generated", flush=True)
            else:
                print(f"  ⚠️ Batch {batch_idx+1}: Empty response, creating placeholders", flush=True)
                for f in batch_features:
                    all_epics.append(_placeholder_epic(f, req.tech_stack))
        except Exception as e:
            print(f"  ⚠️ Batch {batch_idx+1} Error: {e}. Creating placeholders.", flush=True)
            for f in batch_features:
                all_epics.append(_placeholder_epic(f, req.tech_stack))

    parsed["epics"] = all_epics
    print(f"\n✅ Roadmap complete: {len(all_feature_analysis)} features, {len(all_epics)} epics, {sum(len(e.get('stories',[])) for e in all_epics)} stories\n", flush=True)

    return _post_process_roadmap(parsed, req, start_date, target_working_days, target_months, target_sprints, num_features)


def _placeholder_epic(feature, tech_stack):
    """Create a placeholder epic when AI fails for a feature."""
    return {
        "epic_name": feature.get("feature", "Unnamed"),
        "epic_description": feature.get("technical_scope", f"Implementation using {tech_stack}"),
        "feature_type": feature.get("feature_type", "Application (UI)"),
        "total_points": feature.get("story_points", 3),
        "stories": [{
            "summary": f"Implement {feature.get('feature', 'feature')}",
            "description": f"As a user, I want {feature.get('feature', '')} functionality.\n\nAcceptance Criteria:\n- Feature works as specified\n- All edge cases handled\n- QA sign-off complete",
            "story_points": feature.get("story_points", 3),
            "priority": "High"
        }]
    }


def _build_fallback_structure(features, num_features, total_pts, total_seq_days,
                               target_working_days, target_months, start_date):
    """Build a deterministic project structure when AI fails."""
    from datetime import datetime, timedelta

    # Calculate parallel streams needed
    if total_seq_days <= target_working_days:
        streams = 1
        actual_days = total_seq_days
    else:
        streams = min(4, math.ceil(total_seq_days / target_working_days))
        overhead = 1 + (streams - 1) * 0.1
        actual_days = math.ceil((total_seq_days / streams) * overhead)

    actual_months = round(actual_days / 22, 1)
    total_sprints = math.ceil(actual_days / 10)
    velocity = math.ceil(total_pts / max(total_sprints, 1))

    # Simple sequential scheduling
    current_day = 1
    scheduling = []
    for f in features:
        days = f.get("days", 3)
        scheduling.append({
            "id": f.get("id"), "start_day": current_day,
            "end_day": current_day + days - 1,
            "sprint_allocation": f"SP{math.ceil(current_day / 10)}"
        })
        current_day += days

    # Sprint mapping
    sprint_mapping = []
    start_dt = datetime.strptime(start_date, "%Y-%m-%d")
    for sp in range(1, total_sprints + 1):
        sp_start = (sp - 1) * 10 + 1
        sp_end = sp * 10
        sp_date = start_dt + timedelta(days=(sp - 1) * 14)
        sprint_features = [f"Feature #{f.get('id')}" for f in features
                          if scheduling[features.index(f)]["start_day"] >= sp_start
                          and scheduling[features.index(f)]["start_day"] <= sp_end]
        sprint_mapping.append({
            "sprint": f"SP{sp}", "start_day": sp_start, "end_day": sp_end,
            "month": f"M{math.ceil(sp / 2)}", "calendar_month": sp_date.strftime("%b %Y"),
            "features_in_sprint": sprint_features[:5],
            "points_in_sprint": velocity
        })

    # Resource loading
    resource_loading = []
    for d in range(1, actual_days + 1, 10):
        resource_loading.append({
            "day": d, "sprint": f"SP{math.ceil(d / 10)}", "month": f"M{math.ceil(d / 22)}",
            "active_features": min(streams * 2, num_features),
            "team_members_needed": streams * 3 + 2,
            "roles_active": "PM, PP, QA, BA"
        })

    uat_day = actual_days + 1
    pilot_start = uat_day + 5
    pilot_end = pilot_start + 14
    hc_start = pilot_end + 1
    hc_end = hc_start + 14

    end_dt = start_dt + timedelta(days=int(hc_end * 1.4))

    return {
        "sizing_legend": [
            {"size": "XXS", "days": 1, "story_points": 1, "sprints_equivalent": "0.1 sprint"},
            {"size": "SMALL", "days": 2, "story_points": 2, "sprints_equivalent": "0.2 sprint"},
            {"size": "MEDIUM", "days": 3, "story_points": 3, "sprints_equivalent": "0.3 sprint"},
            {"size": "LARGE", "days": 5, "story_points": 5, "sprints_equivalent": "0.5 sprint"},
            {"size": "XL", "days": 8, "story_points": 8, "sprints_equivalent": "0.8 sprint"},
            {"size": "XXL", "days": 13, "story_points": 13, "sprints_equivalent": "1.3 sprints"},
            {"size": "XXXL", "days": 21, "story_points": 21, "sprints_equivalent": "2.1 sprints"},
        ],
        "scheduling": scheduling,
        "team_composition": [
            {"role": "Program Manager (PgM)", "headcount": 1, "billable": True, "justification": "Delivery coordination", "ramp_up_notes": "Day 1", "stream_allocation": "All"},
            {"role": "Tech Lead (TL)", "headcount": streams, "billable": True, "justification": "Technical guidance per stream", "ramp_up_notes": "Day 1", "stream_allocation": f"1 per stream"},
            {"role": "Developer (PP)", "headcount": streams * 2, "billable": True, "justification": "Core development", "ramp_up_notes": "Day 1", "stream_allocation": f"2 per stream"},
            {"role": "QA Engineer", "headcount": streams, "billable": True, "justification": "Testing & quality", "ramp_up_notes": "Day 5", "stream_allocation": f"1 per stream"},
            {"role": "Business Analyst (BA)", "headcount": max(1, streams), "billable": True, "justification": "Requirements & UAT", "ramp_up_notes": "Day 1", "stream_allocation": f"Shared"},
        ],
        "total_team_size": 1 + streams + streams * 2 + streams + max(1, streams),
        "parallel_stream_analysis": {
            "single_stream_days": total_seq_days, "single_stream_months": round(total_seq_days / 22, 1),
            "recommended_streams": streams, "actual_parallel_days": actual_days,
            "actual_parallel_months": actual_months, "target_days": target_working_days,
            "target_months": target_months, "fits_target": actual_days <= target_working_days,
            "coordination_overhead_pct": (streams - 1) * 10,
            "notes": f"Fallback calculation: {streams} streams, {actual_days} working days."
        },
        "timeline": {
            "total_story_points": total_pts, "team_velocity_per_sprint": velocity,
            "sprint_duration_weeks": 2, "total_sprints": total_sprints,
            "total_working_days": actual_days, "total_months": actual_months,
            "start_date": start_date, "end_date": end_dt.strftime("%Y-%m-%d"),
            "assumptions": f"Based on {streams} parallel streams"
        },
        "sprint_mapping": sprint_mapping,
        "gantt_phases": [
            {"phase": "Planning & Discovery", "assigned_roles": "PgM, BA, TL", "dependencies": "None", "start_day": 1, "end_day": 5, "start_week": 1, "end_week": 1, "duration_days": 5, "phase_type": "planning"},
            {"phase": "Development", "assigned_roles": "PP, QA, TL", "dependencies": "Planning", "start_day": 6, "end_day": actual_days, "start_week": 2, "end_week": math.ceil(actual_days / 5), "duration_days": actual_days - 5, "phase_type": "development"},
            {"phase": "Final UAT", "assigned_roles": "QA, BA, PM", "dependencies": "Development", "start_day": uat_day, "end_day": uat_day + 4, "start_week": math.ceil(uat_day / 5), "end_week": math.ceil((uat_day + 4) / 5), "duration_days": 5, "phase_type": "uat"},
            {"phase": "Pilot", "assigned_roles": "PM, PP, QA", "dependencies": "UAT", "start_day": pilot_start, "end_day": pilot_end, "start_week": math.ceil(pilot_start / 5), "end_week": math.ceil(pilot_end / 5), "duration_days": 15, "phase_type": "pilot"},
            {"phase": "Hypercare", "assigned_roles": "PM, PP, QA", "dependencies": "Pilot", "start_day": hc_start, "end_day": hc_end, "start_week": math.ceil(hc_start / 5), "end_week": math.ceil(hc_end / 5), "duration_days": 15, "phase_type": "hypercare"},
        ],
        "resource_loading": resource_loading,
        "uat_milestones": [
            {"name": "UAT Final", "sprint": f"SP{total_sprints + 1}", "day": uat_day, "scope": f"All {num_features} features", "duration_days": 5, "exit_criteria": "All P1 defects resolved, 85% pass rate"}
        ],
        "pilot_hypercare": {
            "pilot": {"start_day": pilot_start, "end_day": pilot_end, "duration_days": 15, "description": "Limited production deployment", "team_needed": "PM, 1 Dev, 1 QA, 1 BA"},
            "hypercare": {"start_day": hc_start, "end_day": hc_end, "duration_days": 15, "description": "Full production support", "team_needed": "PM, 2 Dev, 1 QA"}
        },
        "feature_type_summary": {}
    }


def _post_process_roadmap(parsed, req, start_date, target_working_days, target_months, target_sprints, num_features):
    """
    SERVER-SIDE MATH ENFORCEMENT.
    Treats AI output as a draft — recalculates all derived data deterministically.
    """

    # ═══════════════════════════════════════════════════════════
    # STEP 0: Extract feature analysis (AI sizing is trusted)
    # ═══════════════════════════════════════════════════════════
    fa = parsed.get("feature_analysis", [])

    # Separate dev features from project phases (Pilot/Hypercare)
    dev_features = []
    phase_features = []
    for f in fa:
        name_lower = (f.get("feature", "") or "").lower().strip()
        if name_lower in ("pilot", "hypercare", "pilot phase", "hypercare phase",
                          "pilot & hypercare", "pilot and hypercare"):
            phase_features.append(f)
        else:
            dev_features.append(f)

    total_pts = sum(f.get("story_points", 0) or 0 for f in dev_features)
    total_seq_days = sum(f.get("days", 0) or 0 for f in dev_features)

    parsed["total_story_points"] = total_pts
    parsed["total_working_days_sequential"] = total_seq_days

    # ═══════════════════════════════════════════════════════════
    # STEP 1: Parallel streams (initial read)
    # ═══════════════════════════════════════════════════════════
    psa = parsed.get("parallel_stream_analysis", {})
    if not psa:
        psa = {}
        parsed["parallel_stream_analysis"] = psa

    ai_streams = psa.get("recommended_streams", 1) or 1
    psa["single_stream_days"] = total_seq_days
    psa["single_stream_months"] = round(total_seq_days / 22, 1)

    # ═══════════════════════════════════════════════════════════
    # STEP 2: Count actual developers from team_composition
    # ═══════════════════════════════════════════════════════════
    team = parsed.get("team_composition", [])
    dev_count = 0
    dev_role_keywords = ["developer", "dev ", "engineer", "programmer",
                         "power platform", "bi developer", "frontend",
                         "backend", "full stack", "fullstack"]
    non_dev_keywords = ["manager", "analyst", "qa", "quality", "architect",
                        "scrum", "product owner", "pgm", "program"]

    for member in team:
        role_lower = (member.get("role", "") or "").lower()
        is_dev = any(kw in role_lower for kw in dev_role_keywords)
        is_non_dev = any(kw in role_lower for kw in non_dev_keywords)
        if is_dev and not is_non_dev:
            hc = member.get("headcount", 1) or 1
            dev_count += hc
            print(f"[MATH] DEV: {member.get('role','')} x{hc}", flush=True)

    if dev_count == 0:
        dev_count = max(ai_streams, 1) * 2
        print(f"[MATH] No dev roles detected, defaulting to {dev_count}", flush=True)

    # ═══════════════════════════════════════════════════════════
    # STEP 3: Deterministic velocity
    # ═══════════════════════════════════════════════════════════
    AVG_PTS_PER_DEV_PER_SPRINT = 8
    velocity = dev_count * AVG_PTS_PER_DEV_PER_SPRINT
    if velocity <= 0:
        velocity = 16

    # ═══════════════════════════════════════════════════════════
    # STEP 4: Sprints and duration
    # ═══════════════════════════════════════════════════════════
    total_sprints = math.ceil(total_pts / velocity)
    total_dev_days = total_sprints * 10

    # Validate streams needed
    if total_dev_days <= target_working_days:
        streams = 1
    else:
        streams = min(4, max(1, math.ceil(total_dev_days / target_working_days)))

    psa["recommended_streams"] = streams
    psa["target_days"] = target_working_days
    psa["target_months"] = target_months

    try:
        start_dt = datetime.strptime(start_date, "%Y-%m-%d")
    except Exception:
        start_dt = datetime.now()

    # ═══════════════════════════════════════════════════════════
    # STEP 5: REBUILD sprint_mapping (bin-packing)
    # ═══════════════════════════════════════════════════════════
    sorted_features = sorted(
        dev_features,
        key=lambda f: (f.get("start_day", 999), f.get("id", 999))
    )

    sprint_mapping = []
    feature_queue = list(sorted_features)

    for sp_num in range(1, total_sprints + 1):
        sp_label = f"SP{sp_num}"
        sp_start_day = (sp_num - 1) * 10 + 1
        sp_end_day = sp_num * 10
        sp_date = start_dt + timedelta(days=(sp_num - 1) * 14)
        calendar_month = sp_date.strftime("%b %Y")
        month_label = f"M{math.ceil(sp_num / 2)}"

        sprint_pts = 0
        sprint_features = []
        remaining_queue = []

        for feat in feature_queue:
            feat_pts = feat.get("story_points", 0) or feat.get("days", 3) or 3
            if sprint_pts + feat_pts <= velocity:
                sprint_pts += feat_pts
                feat_name = feat.get("feature", "Unknown")
                if len(feat_name) > 60:
                    feat_name = feat_name[:57] + "..."
                sprint_features.append(feat_name)
                feat["sprint_allocation"] = sp_label
                feat["start_day"] = sp_start_day
                feat["end_day"] = min(sp_start_day + feat_pts - 1, sp_end_day)
            elif sprint_pts == 0 and feat_pts > velocity:
                sprint_pts += feat_pts
                feat_name = feat.get("feature", "Unknown")
                if len(feat_name) > 60:
                    feat_name = feat_name[:57] + "..."
                sprint_features.append(feat_name)
                feat["sprint_allocation"] = sp_label
                feat["start_day"] = sp_start_day
                feat["end_day"] = sp_end_day
                print(f"[MATH] Oversized ({feat_pts}pts) in {sp_label}, cap={velocity}", flush=True)
            else:
                remaining_queue.append(feat)

        feature_queue = remaining_queue
        sprint_mapping.append({
            "sprint": sp_label,
            "start_day": sp_start_day,
            "end_day": sp_end_day,
            "month": month_label,
            "calendar_month": calendar_month,
            "features_in_sprint": sprint_features if sprint_features else ["Buffer"],
            "points_in_sprint": sprint_pts
        })

    # Handle overflow
    while feature_queue:
        total_sprints += 1
        sp_label = f"SP{total_sprints}"
        sp_start_day = (total_sprints - 1) * 10 + 1
        sp_end_day = total_sprints * 10
        sp_date = start_dt + timedelta(days=(total_sprints - 1) * 14)

        sprint_pts = 0
        sprint_features = []
        remaining = []
        for feat in feature_queue:
            feat_pts = feat.get("story_points", 0) or feat.get("days", 3) or 3
            if sprint_pts + feat_pts <= velocity:
                sprint_pts += feat_pts
                feat_name = feat.get("feature", "Unknown")
                sprint_features.append(feat_name[:57] + "..." if len(feat_name) > 60 else feat_name)
                feat["sprint_allocation"] = sp_label
                feat["start_day"] = sp_start_day
                feat["end_day"] = min(sp_start_day + feat_pts - 1, sp_end_day)
            elif sprint_pts == 0:
                sprint_pts += feat_pts
                feat_name = feat.get("feature", "Unknown")
                sprint_features.append(feat_name[:57] + "..." if len(feat_name) > 60 else feat_name)
                feat["sprint_allocation"] = sp_label
            else:
                remaining.append(feat)
        feature_queue = remaining
        sprint_mapping.append({
            "sprint": sp_label,
            "start_day": sp_start_day,
            "end_day": sp_end_day,
            "month": f"M{math.ceil(total_sprints / 2)}",
            "calendar_month": sp_date.strftime("%b %Y"),
            "features_in_sprint": sprint_features,
            "points_in_sprint": sprint_pts
        })
        print(f"[MATH] Overflow {sp_label}: {sprint_pts}pts", flush=True)

    total_dev_days = total_sprints * 10
    parsed["sprint_mapping"] = sprint_mapping

    # ═══════════════════════════════════════════════════════════
    # STEP 6: UAT, Pilot, Hypercare (AFTER sprint count is final)
    # ═══════════════════════════════════════════════════════════
    uat_interval = min(6, max(3, total_sprints // 3))
    uat_sprint_numbers = list(range(uat_interval, total_sprints, uat_interval))
    if not uat_sprint_numbers or uat_sprint_numbers[-1] != total_sprints:
        uat_sprint_numbers.append(total_sprints)

    pilot_days = 15
    hypercare_days = 15
    uat_milestones = []
    current_day = total_dev_days

    for idx, uat_sp in enumerate(uat_sprint_numbers):
        if idx == len(uat_sprint_numbers) - 1:
            uat_milestones.append({
                "name": "Final UAT",
                "sprint": f"SP{uat_sp}",
                "day": current_day + 1,
                "scope": f"All {len(dev_features)} features — full regression",
                "duration_days": 5,
                "exit_criteria": "All P1/P2 resolved, 90% test pass, stakeholder sign-off"
            })
            current_day += 5
        else:
            uat_milestones.append({
                "name": f"UAT {idx + 1}",
                "sprint": f"SP{uat_sp}",
                "day": uat_sp * 10 + 1,
                "scope": f"Features in SP1–SP{uat_sp}",
                "duration_days": 5,
                "exit_criteria": "All P1 resolved, 85% test pass"
            })

    pilot_start = current_day + 1
    pilot_end = pilot_start + pilot_days - 1
    hypercare_start = pilot_end + 1
    hypercare_end = hypercare_start + hypercare_days - 1
    total_project_days = hypercare_end
    total_project_months = round(total_project_days / 22, 1)
    end_dt = start_dt + timedelta(days=int(total_project_days * 1.4))
    end_date_str = end_dt.strftime("%Y-%m-%d")

    parsed["uat_milestones"] = uat_milestones
    parsed["pilot_hypercare"] = {
        "pilot": {"start_day": pilot_start, "end_day": pilot_end,
                  "duration_days": pilot_days,
                  "description": "Limited production deployment with select user group",
                  "team_needed": "PM, Dev Lead, 1 QA, 1 BA"},
        "hypercare": {"start_day": hypercare_start, "end_day": hypercare_end,
                      "duration_days": hypercare_days,
                      "description": "Full production support, monitoring, defect triage",
                      "team_needed": "PM, 2 Dev, 1 QA"}
    }

    # ═══════════════════════════════════════════════════════════
    # STEP 7: REBUILD gantt_phases
    # ═══════════════════════════════════════════════════════════
    gantt_phases = []
    gantt_phases.append({
        "phase": "Planning & Discovery",
        "assigned_roles": "PgM, Architect, BA",
        "dependencies": "None",
        "start_day": 1, "end_day": 5,
        "start_week": 1, "end_week": 1,
        "duration_days": 5,
        "phase_type": "planning"
    })

    dev_phase_num = 1
    dev_start = 1
    for idx, uat in enumerate(uat_milestones):
        if uat["name"].startswith("Final"):
            dev_end = total_dev_days
        else:
            dev_end = uat_sprint_numbers[idx] * 10

        if dev_end >= dev_start:
            gantt_phases.append({
                "phase": f"Development Phase {dev_phase_num}",
                "assigned_roles": "Full Team",
                "dependencies": "Planning" if dev_phase_num == 1 else f"UAT {dev_phase_num - 1}",
                "start_day": dev_start, "end_day": dev_end,
                "start_week": math.ceil(dev_start / 5),
                "end_week": math.ceil(dev_end / 5),
                "duration_days": dev_end - dev_start + 1,
                "phase_type": "development"
            })

        uat_start_day = uat.get("day", dev_end + 1)
        gantt_phases.append({
            "phase": uat["name"],
            "assigned_roles": "QA, BA, Business Users",
            "dependencies": f"Development Phase {dev_phase_num}",
            "start_day": uat_start_day, "end_day": uat_start_day + 4,
            "start_week": math.ceil(uat_start_day / 5),
            "end_week": math.ceil((uat_start_day + 4) / 5),
            "duration_days": 5,
            "phase_type": "uat"
        })
        dev_start = dev_end + 1 if not uat["name"].startswith("Final") else dev_end + 6
        dev_phase_num += 1

    gantt_phases.append({
        "phase": "Pilot Deployment", "assigned_roles": "PM, Dev Lead, QA, BA",
        "dependencies": "Final UAT",
        "start_day": pilot_start, "end_day": pilot_end,
        "start_week": math.ceil(pilot_start / 5), "end_week": math.ceil(pilot_end / 5),
        "duration_days": pilot_days, "phase_type": "pilot"
    })
    gantt_phases.append({
        "phase": "Hypercare & Stabilization", "assigned_roles": "PM, Dev, QA",
        "dependencies": "Pilot",
        "start_day": hypercare_start, "end_day": hypercare_end,
        "start_week": math.ceil(hypercare_start / 5), "end_week": math.ceil(hypercare_end / 5),
        "duration_days": hypercare_days, "phase_type": "hypercare"
    })
    parsed["gantt_phases"] = gantt_phases

    # ═══════════════════════════════════════════════════════════
    # STEP 8: REBUILD resource_loading
    # ═══════════════════════════════════════════════════════════
    total_team_size = sum(m.get("headcount", 1) or 1 for m in team) if team else (dev_count + 3)
    resource_loading = []
    for d in range(1, total_project_days + 1, 10):
        sp_num = math.ceil(d / 10)
        month_num = math.ceil(d / 22)
        if d <= total_dev_days:
            active = 0
            for sm in sprint_mapping:
                if sm["start_day"] <= d <= sm["end_day"]:
                    active = len(sm.get("features_in_sprint", []))
                    break
            resource_loading.append({"day": d, "sprint": f"SP{sp_num}", "month": f"M{month_num}",
                                     "active_features": active, "team_members_needed": total_team_size,
                                     "roles_active": "PgM, Dev, QA, BA"})
        elif d <= total_dev_days + 5:
            resource_loading.append({"day": d, "sprint": "UAT", "month": f"M{month_num}",
                                     "active_features": 0, "team_members_needed": max(3, total_team_size - dev_count + 1),
                                     "roles_active": "QA, BA, PM"})
        elif d <= pilot_end:
            resource_loading.append({"day": d, "sprint": "Pilot", "month": f"M{month_num}",
                                     "active_features": 0, "team_members_needed": 4,
                                     "roles_active": "PM, Dev Lead, QA, BA"})
        else:
            resource_loading.append({"day": d, "sprint": "Hypercare", "month": f"M{month_num}",
                                     "active_features": 0, "team_members_needed": 3,
                                     "roles_active": "PM, Dev, QA"})
    parsed["resource_loading"] = resource_loading

    # ═══════════════════════════════════════════════════════════
    # STEP 9: Fix timeline
    # ═══════════════════════════════════════════════════════════
    tl = parsed.get("timeline", {})
    if not tl:
        tl = {}
        parsed["timeline"] = tl

    tl["total_story_points"] = total_pts
    tl["team_velocity_per_sprint"] = velocity
    tl["sprint_duration_weeks"] = 2
    tl["total_sprints"] = total_sprints
    tl["total_working_days"] = total_dev_days
    tl["total_months"] = round(total_dev_days / 22, 1)
    tl["start_date"] = start_date
    tl["end_date"] = end_date_str
    tl["assumptions"] = (
        f"{dev_count} developers × {AVG_PTS_PER_DEV_PER_SPRINT} pts/dev/sprint = "
        f"{velocity} velocity. {total_sprints} dev sprints + "
        f"{len(uat_milestones)} UAT gates + Pilot ({pilot_days}d) + Hypercare ({hypercare_days}d). "
        f"Total project: {total_project_days} working days ({total_project_months} months)."
    )

    # ═══════════════════════════════════════════════════════════
    # STEP 10: Fix parallel_stream_analysis
    # ═══════════════════════════════════════════════════════════
    psa["actual_parallel_days"] = total_dev_days
    psa["actual_parallel_months"] = round(total_dev_days / 22, 1)
    psa["fits_target"] = total_dev_days <= target_working_days
    psa["coordination_overhead_pct"] = (streams - 1) * 10
    psa["notes"] = (
        f"{dev_count} devs × {AVG_PTS_PER_DEV_PER_SPRINT} pts = {velocity} pts/sprint. "
        f"{total_sprints} sprints ({total_dev_days} days) for {total_pts} story points. "
        f"{'Fits' if total_dev_days <= target_working_days else 'Exceeds'} "
        f"the {target_working_days}-day ({target_months}m) target."
    )

    # ═══════════════════════════════════════════════════════════
    # STEP 11: Ensure epics & feature_type_summary
    # ═══════════════════════════════════════════════════════════
    if "epics" not in parsed:
        parsed["epics"] = []
    if "feature_type_summary" not in parsed:
        ft_summary = {}
        for f in fa:
            ft = f.get("feature_type", "Application (UI)")
            if ft not in ft_summary:
                ft_summary[ft] = {"count": 0, "total_points": 0}
            ft_summary[ft]["count"] += 1
            ft_summary[ft]["total_points"] += f.get("story_points", 0)
        parsed["feature_type_summary"] = ft_summary

    # ═══════════════════════════════════════════════════════════
    # STEP 12: Metadata + Math Audit
    # ═══════════════════════════════════════════════════════════
    sprint_violations = []
    sprint_point_sum = 0
    for sm in sprint_mapping:
        pts = sm.get("points_in_sprint", 0)
        sprint_point_sum += pts
        if pts > velocity * 1.1:
            sprint_violations.append(f"{sm['sprint']}: {pts}pts (cap:{velocity})")

    parsed["_meta"] = {
        "tech_stack": req.tech_stack,
        "target_duration_value": req.target_duration_value,
        "target_duration_unit": req.target_duration_unit,
        "target_working_days": target_working_days,
        "target_months": target_months,
        "target_sprints": target_sprints,
        "num_features": num_features,
        "num_dev_features": len(dev_features),
        "num_phase_features_excluded": len(phase_features),
        "start_date": start_date,
        "generated_at": datetime.now().isoformat(),
        "math_audit": {
            "total_story_points": total_pts,
            "total_sequential_days": total_seq_days,
            "parallel_streams": streams,
            "developers_detected": dev_count,
            "velocity_formula": f"{dev_count} devs × {AVG_PTS_PER_DEV_PER_SPRINT} pts = {velocity} pts/sprint",
            "sprints_formula": f"ceil({total_pts}/{velocity}) = {total_sprints}",
            "total_dev_sprints": total_sprints,
            "total_dev_days": total_dev_days,
            "uat_gates": len(uat_milestones),
            "pilot_days": pilot_days,
            "hypercare_days": hypercare_days,
            "total_project_days": total_project_days,
            "total_project_months": total_project_months,
            "sprint_point_sum": sprint_point_sum,
            "sprint_violations": sprint_violations if sprint_violations else "None",
            "end_date": end_date_str
        }
    }

    print(f"\n{'='*60}", flush=True)
    print(f"📐 MATH ENFORCEMENT COMPLETE", flush=True)
    print(f"   Features: {len(dev_features)} dev + {len(phase_features)} phase", flush=True)
    print(f"   Points: {total_pts} | Seq Days: {total_seq_days}", flush=True)
    print(f"   Devs: {dev_count} | Velocity: {velocity} pts/sprint", flush=True)
    print(f"   Sprints: {total_sprints} | Dev Days: {total_dev_days}", flush=True)
    print(f"   Project: {total_project_days}d ({total_project_months}m)", flush=True)
    print(f"   Sprint Sum: {sprint_point_sum} | Violations: {len(sprint_violations)}", flush=True)
    print(f"{'='*60}\n", flush=True)

    return parsed


def _build_full_roadmap_prompt(tech_stack, features_text, num_features, start_date,
                                target_working_days, target_months, target_sprints):
    """Build the full single-call prompt (for ≤20 features)."""
    return f"""You are a world-class Technical Project Manager, Solutions Architect, and Agile Delivery Lead with 20+ years of experience estimating enterprise software projects.

CONTEXT:
- Tech Stack: {tech_stack}
- Number of Features: {num_features}
- Client Target Deadline: {target_working_days} working days ({target_months} months, ~{target_sprints} sprints)
- Project Start Date: {start_date}

CLIENT FEATURES TO BUILD:
{features_text}

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
CRITICAL: DETERMINISTIC VELOCITY & DURATION MATH
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
This data will be shared directly with the Client. ALL math must be verifiable.

STEP 1 — Total Story Points (TSP): Sum ALL feature story points. FIXED number.
STEP 2 — Total Sequential Days (TSD): Sum ALL feature days. days = story_points.
STEP 3 — Parallel Streams: streams = ceil(TSD / {target_working_days}). Max 4. +10% overhead per extra stream.
STEP 4 — Team per Stream: 1 TL + 2 PP + 1 QA + 1 BA. Plus 1 PgM cross-stream.
STEP 5 — Velocity: team_velocity_per_sprint = (streams × 2 devs) × 8 pts/dev/sprint
  HARD RULE: No sprint can exceed team_velocity_per_sprint.
STEP 6 — Duration: total_sprints = ceil(TSP / velocity). total_days = sprints × 10.
  RULE: If 1 stream → total_working_days MUST = TSD.

VALIDATION (verify before responding):
□ sum(feature points) == total_story_points == timeline.total_story_points
□ sum(feature days) == total_working_days_sequential == parallel_stream_analysis.single_stream_days
□ velocity == (streams × 2) × 8
□ Every sprint: points_in_sprint <= velocity
□ sum(sprint points) == total_story_points
□ If 1 stream: total_working_days == sequential days

EXAMPLE: 10 features, 65 pts, target 44 days
→ streams = ceil(65/44) = 2 → devs = 4 → velocity = 32/sprint
→ sprints = ceil(65/32) = 3 → days = 30 → with overhead = 33 ✓

DO NOT: Invent velocity, show inconsistent totals, exceed sprint capacity, or deviate from formula.
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

Analyze ALL features and return ONLY valid JSON (no markdown, no backticks) with this EXACT structure:

{{
    "sizing_legend": [
        {{"size": "XXS", "days": 1, "story_points": 1, "sprints_equivalent": "0.1 sprint"}},
        {{"size": "SMALL", "days": 2, "story_points": 2, "sprints_equivalent": "0.2 sprint"}},
        {{"size": "MEDIUM", "days": 3, "story_points": 3, "sprints_equivalent": "0.3 sprint"}},
        {{"size": "LARGE", "days": 5, "story_points": 5, "sprints_equivalent": "0.5 sprint"}},
        {{"size": "XL", "days": 8, "story_points": 8, "sprints_equivalent": "0.8 sprint"}},
        {{"size": "XXL", "days": 13, "story_points": 13, "sprints_equivalent": "1.3 sprints"}},
        {{"size": "XXXL", "days": 21, "story_points": 21, "sprints_equivalent": "2.1 sprints"}}
    ],
    "feature_analysis": [
        {{
            "id": 1,
            "feature": "Exact feature name from input",
            "feature_type": "Application (UI)",
            "technical_scope": "Detailed technical scope based on {tech_stack} (2-3 sentences)",
            "size": "XXL",
            "story_points": 13,
            "days": 13,
            "roles_needed": "PP, QA, BA, INT",
            "est_team": "XXL – PP, QA, BA, INT",
            "est_conservative": "L – PP, QA, BA",
            "dependencies": "Feature #3, Feature #5",
            "start_day": 1,
            "end_day": 13,
            "sprint_allocation": "SP1-SP2"
        }}
    ],
    "total_story_points": 0,
    "total_working_days_sequential": 0,
    "team_composition": [
        {{
            "role": "Program Manager (PgM)",
            "headcount": 1,
            "billable": true,
            "justification": "Overall delivery coordination",
            "ramp_up_notes": "From Day 1",
            "stream_allocation": "All streams"
        }}
    ],
    "total_team_size": 0,
    "parallel_stream_analysis": {{
        "single_stream_days": 0,
        "single_stream_months": 0,
        "recommended_streams": 0,
        "actual_parallel_days": 0,
        "actual_parallel_months": 0,
        "target_days": {target_working_days},
        "target_months": {target_months},
        "fits_target": true,
        "coordination_overhead_pct": 10,
        "notes": "With N parallel streams, the project fits within the target of {target_months} months. Team size scaled accordingly."
    }},
    "timeline": {{
        "total_story_points": 0,
        "team_velocity_per_sprint": 0,
        "sprint_duration_weeks": 2,
        "total_sprints": 0,
        "total_working_days": 0,
        "total_months": 0,
        "start_date": "{start_date}",
        "end_date": "YYYY-MM-DD",
        "assumptions": "Based on N parallel streams with team of X"
    }},
    "sprint_mapping": [
        {{
            "sprint": "SP1",
            "start_day": 1,
            "end_day": 10,
            "month": "M1",
            "calendar_month": "Mar 2026",
            "features_in_sprint": ["Feature #1", "Feature #2"],
            "points_in_sprint": 25
        }}
    ],
    "gantt_phases": [
        {{
            "phase": "Planning & Discovery",
            "assigned_roles": "PgM, Architect, Tech Lead",
            "dependencies": "None",
            "start_day": 1,
            "end_day": 5,
            "start_week": 1,
            "end_week": 1,
            "duration_days": 5,
            "phase_type": "planning"
        }}
    ],
    "resource_loading": [
        {{
            "day": 1,
            "sprint": "SP1",
            "month": "M1",
            "active_features": 2,
            "team_members_needed": 4,
            "roles_active": "PP, QA, BA, PM"
        }}
    ],
    "uat_milestones": [
        {{
            "name": "UAT 1",
            "sprint": "SP8",
            "day": 80,
            "scope": "Core modules: Auth, Dashboard, Client Management",
            "duration_days": 5,
            "exit_criteria": "All P1 defects resolved, 85% test pass rate"
        }}
    ],
    "pilot_hypercare": {{
        "pilot": {{
            "start_day": 0,
            "end_day": 0,
            "duration_days": 15,
            "description": "Limited production deployment with select users",
            "team_needed": "PM, 1 Dev, 1 QA, 1 BA"
        }},
        "hypercare": {{
            "start_day": 0,
            "end_day": 0,
            "duration_days": 15,
            "description": "Full production support, monitoring, defect resolution",
            "team_needed": "PM, 2 Dev, 1 QA"
        }}
    }},
    "epics": [
        {{
            "epic_name": "User Authentication & Authorization",
            "epic_description": "All features related to user access control",
            "feature_type": "Application (UI)",
            "total_points": 13,
            "stories": [
                {{
                    "summary": "Implement user registration with email verification",
                    "description": "As a new user, I want to register with my email so I can access the platform.\\n\\nAcceptance Criteria:\\n- Email + password registration form\\n- Email verification flow\\n- Error handling for duplicate emails",
                    "story_points": 5,
                    "priority": "High"
                }}
            ]
        }}
    ],
    "feature_type_summary": {{
        "Application (UI)": {{"count": 0, "total_points": 0}},
        "Systems Integration": {{"count": 0, "total_points": 0}},
        "Reporting": {{"count": 0, "total_points": 0}},
        "Process Automation": {{"count": 0, "total_points": 0}},
        "Data & Backend": {{"count": 0, "total_points": 0}}
    }}
}}

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ESTIMATION RULES:
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
1. T-Shirt sizes: XXS=1, SMALL=2, MEDIUM=3, LARGE=5, XL=8, XXL=13, XXXL=21 (Fibonacci ONLY)
2. Each feature gets: size, points, days (same as points), roles_needed, dual estimation
3. Feature types: "Application (UI)", "Systems Integration", "Reporting", "Process Automation", "Data & Backend"
4. est_team = your team-level estimate with roles (e.g. "XXL – PP, QA, BA, INT")
5. est_conservative = a more conservative/TA estimate (often 1 size smaller)
6. Per-feature scheduling: start_day and end_day for EVERY feature (day-level Gantt)
7. Features with dependencies start AFTER their dependency ends
8. Parallel features share the same day range ONLY if role assignments don't conflict
9. Resource loading: for every 10th day, show team members needed
10. UAT gates: place UAT checkpoints every ~4-6 sprints, plus a FINAL UAT before pilot
11. Pilot: 10-15 working days after all development + final UAT
12. Hypercare: 10-15 working days after pilot
13. Sprint mapping: SP1 starts Day 1-10, SP2 Day 11-20, etc.
14. All {num_features} features MUST be present in feature_analysis
15. Calendar months should start from {start_date}
16. Ensure resource_loading has entries for day 1, 10, 20, 30... up to total_working_days
17. gantt_phases must include: Planning, Setup/CI-CD, Development (grouped by sprint), Testing/QA, each UAT, Pilot, Hypercare
18. Parallel stream analysis must show: what if 1 stream vs recommended streams"""
   
@app.post("/feature_roadmap/push_jira/{project_key}")
def push_feature_roadmap_to_jira(project_key: str, req: JiraPushRequest, creds: dict = Depends(get_jira_creds)):
    sp_field = get_story_point_field(creds)
    epic_name_field = get_epic_name_field(creds)  # Discover dynamically
    created = []
    errors = []

    # Get base URL for issue links
    base_url = ""
    if creds.get("auth_type") == "basic":
        base_url = f"https://{creds['domain']}"
    else:
        server_info = jira_request("GET", "serverInfo", creds)
        if server_info is not None and server_info.status_code == 200:
            base_url = server_info.json().get("baseUrl", "")

    print(f"\n{'='*60}", flush=True)
    print(f"📤 PUSH TO JIRA: {len(req.epics)} epics for {project_key}", flush=True)
    print(f"   SP Field: {sp_field} | Epic Name Field: {epic_name_field}", flush=True)
    print(f"{'='*60}\n", flush=True)

    for epic_idx, epic in enumerate(req.epics):
        epic_name = epic.get("epic_name", "Unnamed Epic")
        epic_desc = epic.get("epic_description", "")

        print(f"  [{epic_idx+1}/{len(req.epics)}] Creating Epic: {epic_name}", flush=True)

        epic_payload = {
            "fields": {
                "project": {"key": project_key},
                "summary": epic_name,
                "description": {
                    "type": "doc", "version": 1,
                    "content": [{"type": "paragraph", "content": [{"type": "text", "text": epic_desc or "AI Generated Epic"}]}]
                },
                "issuetype": {"name": "Epic"},
            }
        }

        # Only set Epic Name field if we discovered it AND it's not the SP field
        if epic_name_field and epic_name_field != sp_field:
            epic_payload["fields"][epic_name_field] = epic_name

        # ═══ ATTEMPT EPIC CREATION WITH FALLBACK ═══
        epic_res = jira_request("POST", "issue", creds, epic_payload)

        # If Epic type fails, try without Epic Name field
        if (not epic_res or epic_res.status_code not in [200, 201]) and epic_name_field:
            print(f"    ⚠️ Epic creation failed with epic name field, retrying without it...", flush=True)
            epic_payload["fields"].pop(epic_name_field, None)
            epic_res = jira_request("POST", "issue", creds, epic_payload)

        # If Epic type doesn't exist, fall back to Story
        if not epic_res or epic_res.status_code not in [200, 201]:
            print(f"    ⚠️ Epic type failed, falling back to Story type...", flush=True)
            epic_payload["fields"]["issuetype"]["name"] = "Story"
            epic_payload["fields"].pop(epic_name_field, None)  # Story doesn't need epic name
            epic_res = jira_request("POST", "issue", creds, epic_payload)

        # Last resort: Task
        if not epic_res or epic_res.status_code not in [200, 201]:
            print(f"    ⚠️ Story type failed, falling back to Task type...", flush=True)
            epic_payload["fields"]["issuetype"]["name"] = "Task"
            epic_res = jira_request("POST", "issue", creds, epic_payload)

        if not epic_res or epic_res.status_code not in [200, 201]:
            error_detail = extract_jira_error(epic_res)
            print(f"    ❌ FAILED to create epic '{epic_name}': {error_detail}", flush=True)
            errors.append({"epic": epic_name, "error": f"Failed to create: {error_detail}"})
            continue

        epic_key = epic_res.json().get("key", "UNKNOWN")
        epic_id = epic_res.json().get("id", "")
        actual_type = epic_payload["fields"]["issuetype"]["name"]
        epic_url = f"{base_url}/browse/{epic_key}" if base_url else ""
        print(f"    ✅ Created {actual_type}: {epic_key}", flush=True)

        created_stories = []

        for story_idx, story in enumerate(epic.get("stories", [])):
            story_summary = story.get("summary", "Untitled Story")
            print(f"      [{story_idx+1}] Creating Story: {story_summary[:60]}...", flush=True)

            story_desc_text = story.get("description", "")
            desc_content = []
            for paragraph in story_desc_text.split("\n"):
                if paragraph.strip():
                    desc_content.append({
                        "type": "paragraph",
                        "content": [{"type": "text", "text": paragraph.strip()}]
                    })

            story_payload = {
                "fields": {
                    "project": {"key": project_key},
                    "summary": story_summary,
                    "description": {
                        "type": "doc", "version": 1,
                        "content": desc_content if desc_content else [
                            {"type": "paragraph", "content": [{"type": "text", "text": "No description"}]}
                        ]
                    },
                    "issuetype": {"name": "Story"},
                    "priority": {"name": story.get("priority", "Medium")},
                }
            }

            # Link to parent Epic (only if parent was actually created as Epic)
            if epic_id and actual_type == "Epic":
                story_payload["fields"]["parent"] = {"id": epic_id}

            # Set story points ONLY if field is known and value is valid
            pts = safe_float(story.get("story_points", 0))
            if pts > 0 and sp_field:
                story_payload["fields"][sp_field] = pts

            story_res = jira_request("POST", "issue", creds, story_payload)

            # If Story fails with SP field, retry without it
            if (not story_res or story_res.status_code not in [200, 201]) and pts > 0:
                print(f"        ⚠️ Story creation failed, retrying without SP in payload...", flush=True)
                story_payload["fields"].pop(sp_field, None)
                story_res = jira_request("POST", "issue", creds, story_payload)

            # Fallback to Task
            if not story_res or story_res.status_code not in [200, 201]:
                print(f"        ⚠️ Story type failed, falling back to Task...", flush=True)
                story_payload["fields"]["issuetype"]["name"] = "Task"
                story_payload["fields"].pop("parent", None)  # Task may not support parent
                story_res = jira_request("POST", "issue", creds, story_payload)

            if story_res and story_res.status_code in [200, 201]:
                story_key = story_res.json().get("key", "UNKNOWN")
                print(f"        ✅ Created: {story_key}", flush=True)

                # Set story points via separate PUT if not set in create
                if pts > 0 and sp_field not in story_payload.get("fields", {}):
                    pts_res = jira_request("PUT", f"issue/{story_key}", creds, {"fields": {sp_field: pts}})
                    if pts_res and pts_res.status_code in [200, 204]:
                        print(f"        📊 Set {pts} pts on {story_key}", flush=True)

                story_url = f"{base_url}/browse/{story_key}" if base_url else ""
                created_stories.append({
                    "key": story_key,
                    "summary": story_summary,
                    "points": pts,
                    "url": story_url
                })
            else:
                error_detail = extract_jira_error(story_res)
                print(f"        ❌ FAILED: {story_summary[:50]} — {error_detail}", flush=True)
                errors.append({
                    "epic": epic_name,
                    "story": story_summary,
                    "error": f"Failed: {error_detail}"
                })

        created.append({
            "epic_key": epic_key,
            "epic_name": epic_name,
            "epic_url": epic_url,
            "stories_created": len(created_stories),
            "stories": created_stories
        })

    total_epics = len(created)
    total_stories = sum(e["stories_created"] for e in created)
    print(f"\n{'='*60}", flush=True)
    print(f"📤 PUSH COMPLETE: {total_epics} epics, {total_stories} stories, {len(errors)} errors", flush=True)
    print(f"{'='*60}\n", flush=True)

    return {
        "success": total_epics > 0 or total_stories > 0,
        "created_epics": total_epics,
        "created_stories": total_stories,
        "details": created,
        "errors": errors
    }

EPIC_NAME_FIELD_CACHE = {}

def get_epic_name_field(creds):
    """Discover the correct Epic Name field ID for this Jira instance."""
    domain_key = creds.get('domain') or creds.get('cloud_id')
    if domain_key in EPIC_NAME_FIELD_CACHE:
        return EPIC_NAME_FIELD_CACHE[domain_key]

    res = jira_request("GET", "field", creds)
    if res is not None and res.status_code == 200:
        fields = res.json()
        # Look for Epic Name field specifically
        for f in fields:
            name_lower = f.get('name', '').lower()
            if name_lower == 'epic name':
                EPIC_NAME_FIELD_CACHE[domain_key] = f['id']
                print(f"[PUSH] Discovered Epic Name field: {f['id']}", flush=True)
                return f['id']
        # Fallback: look for Epic Link or similar
        for f in fields:
            if 'epic' in f.get('name', '').lower() and 'name' in f.get('name', '').lower():
                EPIC_NAME_FIELD_CACHE[domain_key] = f['id']
                print(f"[PUSH] Discovered Epic Name field (fallback): {f['id']}", flush=True)
                return f['id']

    # Return None if not found — we'll skip setting it
    EPIC_NAME_FIELD_CACHE[domain_key] = None
    return None

@app.post("/feature_roadmap/download_xlsx")
def download_feature_roadmap_xlsx(req: XLSXDownloadRequest):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # ── Safe extraction helpers ──
    def safe_str(obj, key, default=""):
        if not isinstance(obj, dict): return str(default)
        val = obj.get(key)
        if val is None: return str(default)
        return str(val)

    def safe_int(obj, key, default=0):
        if not isinstance(obj, dict): return default
        val = obj.get(key)
        if val is None: return default
        try: return int(float(val))
        except (ValueError, TypeError): return default

    def safe_num(obj, key, default=0):
        if not isinstance(obj, dict): return default
        val = obj.get(key)
        if val is None: return default
        try: return float(val)
        except (ValueError, TypeError): return default

    def safe_list(obj, key):
        if not isinstance(obj, dict): return []
        val = obj.get(key)
        if isinstance(val, list): return val
        return []

    def safe_dict(obj, key):
        if not isinstance(obj, dict): return {}
        val = obj.get(key)
        if isinstance(val, dict): return val
        return {}

    def safe_bool(obj, key, default=True):
        if not isinstance(obj, dict): return default
        val = obj.get(key)
        if val is None: return default
        return bool(val)

    try:
        data = req.data
        if not isinstance(data, dict):
            print(f"[XLSX] ERROR: req.data is not a dict, it's {type(data)}", flush=True)
            raise HTTPException(status_code=400, detail="Invalid data format — expected JSON object")

        print(f"[XLSX] Starting generation. Keys in data: {list(data.keys())}", flush=True)

        # ── Extract all sections with safe defaults ──
        meta = safe_dict(data, "_meta")
        fa = safe_list(data, "feature_analysis")
        team = safe_list(data, "team_composition")
        tl = safe_dict(data, "timeline")
        gantt = safe_list(data, "gantt_phases")
        epics = safe_list(data, "epics")
        sprint_map = safe_list(data, "sprint_mapping")
        resource_load = safe_list(data, "resource_loading")
        uat = safe_list(data, "uat_milestones")
        ph = safe_dict(data, "pilot_hypercare")
        psa = safe_dict(data, "parallel_stream_analysis")
        sizing = safe_list(data, "sizing_legend")

        pilot = safe_dict(ph, "pilot")
        hypercare = safe_dict(ph, "hypercare")

        print(f"[XLSX] Data loaded: {len(fa)} features, {len(team)} team, {len(epics)} epics, {len(sprint_map)} sprints, {len(gantt)} phases", flush=True)

        wb = Workbook()

        # ── Styles ──
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill("solid", fgColor="2563EB")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        size_colors = {
            'XXS': 'E0F7FA', 'SMALL': 'E8F5E9', 'MEDIUM': 'E3F2FD',
            'LARGE': 'FFF8E1', 'XL': 'FFF3E0', 'XXL': 'FCE4EC', 'XXXL': 'FFEBEE'
        }

        def write_header(ws, row, headers, col_start=1):
            for i, h in enumerate(headers):
                cell = ws.cell(row=row, column=col_start + i, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', wrap_text=True)
                cell.border = thin_border

        def auto_width(ws, min_w=10, max_w=40):
            try:
                for col_cells in ws.columns:
                    max_len = 0
                    col_letter = get_column_letter(col_cells[0].column)
                    for cell in col_cells:
                        if cell.value:
                            max_len = max(max_len, len(str(cell.value)))
                    ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_w), max_w)
            except Exception as e:
                print(f"[XLSX] auto_width warning: {e}", flush=True)

        # ═══════════════════════════════════════════════
        # SHEET 1: Project Summary
        # ═══════════════════════════════════════════════
        try:
            print("[XLSX] Building Sheet 1: Project Summary...", flush=True)
            ws1 = wb.active
            ws1.title = "Project Summary"
            ws1.sheet_properties.tabColor = "2563EB"

            ws1['A1'] = "FEATURE ROADMAP — PROJECT ESTIMATION"
            ws1['A1'].font = Font(bold=True, size=16, color="2563EB")
            ws1.merge_cells('A1:F1')
            ws1['A2'] = f"Tech Stack: {safe_str(meta, 'tech_stack', 'N/A')} | Target: {safe_str(meta, 'target_duration_value', '')} {safe_str(meta, 'target_duration_unit', '')} | Generated: {safe_str(meta, 'generated_at', '')[:10]}"
            ws1['A2'].font = Font(italic=True, size=10, color="666666")
            ws1.merge_cells('A2:F2')

            # Sizing Legend
            r = 4
            ws1.cell(row=r, column=1, value="T-SHIRT SIZING LEGEND").font = Font(bold=True, size=12, color="2563EB")
            r += 1
            write_header(ws1, r, ["Size", "Days", "Story Points", "Sprint Equivalent"])
            r += 1
            for s in sizing:
                if not isinstance(s, dict): continue
                sz = safe_str(s, "size")
                ws1.cell(row=r, column=1, value=sz).border = thin_border
                ws1.cell(row=r, column=1).fill = PatternFill("solid", fgColor=size_colors.get(sz, "FFFFFF"))
                ws1.cell(row=r, column=2, value=safe_int(s, "days")).border = thin_border
                ws1.cell(row=r, column=3, value=safe_int(s, "story_points")).border = thin_border
                ws1.cell(row=r, column=4, value=safe_str(s, "sprints_equivalent")).border = thin_border
                r += 1

            # Team Composition
            r += 1
            ws1.cell(row=r, column=1, value="TEAM COMPOSITION").font = Font(bold=True, size=12, color="2563EB")
            r += 1
            write_header(ws1, r, ["Role", "Headcount", "Billable", "Justification", "Ramp-Up Notes", "Stream Allocation"])
            r += 1
            for t in team:
                if not isinstance(t, dict): continue
                ws1.cell(row=r, column=1, value=safe_str(t, "role")).border = thin_border
                ws1.cell(row=r, column=2, value=safe_int(t, "headcount", 1)).border = thin_border
                ws1.cell(row=r, column=3, value="Yes" if safe_bool(t, "billable") else "No").border = thin_border
                ws1.cell(row=r, column=4, value=safe_str(t, "justification")).border = thin_border
                ws1.cell(row=r, column=5, value=safe_str(t, "ramp_up_notes")).border = thin_border
                ws1.cell(row=r, column=6, value=safe_str(t, "stream_allocation")).border = thin_border
                r += 1
            total_hc = sum(safe_int(t, "headcount", 1) for t in team if isinstance(t, dict))
            ws1.cell(row=r, column=1, value="TOTAL TEAM SIZE").font = Font(bold=True)
            ws1.cell(row=r, column=2, value=safe_int(data, "total_team_size", total_hc)).font = Font(bold=True, color="2563EB", size=14)

            # Parallel Stream Analysis
            r += 2
            ws1.cell(row=r, column=1, value="PARALLEL STREAM ANALYSIS").font = Font(bold=True, size=12, color="2563EB")
            r += 1
            for label, key in [
                ("Sequential (1 Stream) Days", "single_stream_days"),
                ("Sequential (1 Stream) Months", "single_stream_months"),
                ("Recommended Parallel Streams", "recommended_streams"),
                ("Actual Duration (Parallel) Days", "actual_parallel_days"),
                ("Actual Duration (Parallel) Months", "actual_parallel_months"),
                ("Client Target Days", "target_days"),
                ("Client Target Months", "target_months"),
                ("Fits Target?", "fits_target"),
                ("Coordination Overhead", "coordination_overhead_pct"),
            ]:
                ws1.cell(row=r, column=1, value=label).font = Font(bold=True)
                ws1.cell(row=r, column=1).border = thin_border
                val = psa.get(key, "")
                if key == "fits_target":
                    display_val = "YES" if val else "NO"
                elif key == "coordination_overhead_pct":
                    display_val = f"{val}%"
                else:
                    display_val = str(val) if val is not None else ""
                ws1.cell(row=r, column=2, value=display_val).border = thin_border
                r += 1
            notes = safe_str(psa, "notes")
            if notes:
                ws1.cell(row=r, column=1, value="Notes:").font = Font(italic=True)
                ws1.cell(row=r, column=2, value=notes)
                ws1.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)

            # Timeline Summary
            r += 2
            ws1.cell(row=r, column=1, value="TIMELINE SUMMARY").font = Font(bold=True, size=12, color="2563EB")
            r += 1
            tl_items = [
                ("Total Story Points", safe_str(tl, "total_story_points", str(data.get("total_story_points", 0)))),
                ("Team Velocity/Sprint", safe_str(tl, "team_velocity_per_sprint")),
                ("Sprint Duration", "2 weeks"),
                ("Total Sprints", safe_str(tl, "total_sprints")),
                ("Total Working Days", safe_str(tl, "total_working_days")),
                ("Total Calendar Months", safe_str(tl, "total_months")),
                ("Start Date", safe_str(tl, "start_date")),
                ("End Date", safe_str(tl, "end_date")),
                ("Assumptions", safe_str(tl, "assumptions")),
            ]
            for label, val in tl_items:
                ws1.cell(row=r, column=1, value=label).font = Font(bold=True)
                ws1.cell(row=r, column=1).border = thin_border
                ws1.cell(row=r, column=2, value=val).border = thin_border
                r += 1

            auto_width(ws1)
            print("[XLSX] Sheet 1 done.", flush=True)
        except Exception as e:
            print(f"[XLSX] Sheet 1 ERROR: {e}", flush=True)
            traceback.print_exc()

        # ═══════════════════════════════════════════════
        # SHEET 2: Feature Schedule (Gantt)
        # ═══════════════════════════════════════════════
        try:
            print("[XLSX] Building Sheet 2: Feature Schedule...", flush=True)
            ws2 = wb.create_sheet("Feature Schedule")
            ws2.sheet_properties.tabColor = "10B981"

            # Calculate total days safely
            all_end_days = [safe_int(f, "end_day", 0) for f in fa if isinstance(f, dict)]
            hc_end = safe_int(hypercare, "end_day", 0)
            all_end_days.append(hc_end)
            total_days = max(all_end_days) if all_end_days else 10
            total_days = max(total_days, 10)
            print(f"[XLSX] Total days for Gantt: {total_days}", flush=True)

            static_cols = ["ID", "Feature Name", "Type", "Days", "Sizing", "Est. TEAM", "Est. Conservative", "Roles", "Dependencies"]
            num_static = len(static_cols)
            write_header(ws2, 1, static_cols)

            # Sprint headers
            for d in range(1, total_days + 1):
                col = num_static + d
                sprint_num = math.ceil(d / 10)
                if (d - 1) % 10 == 0:
                    cell = ws2.cell(row=1, column=col, value=f"SP{sprint_num}")
                    cell.font = Font(bold=True, size=8, color="2563EB")
                    cell.alignment = Alignment(horizontal='center')

            # Day number row
            for d in range(1, total_days + 1):
                cell = ws2.cell(row=2, column=num_static + d, value=d)
                cell.font = Font(size=7, color="999999")
                cell.alignment = Alignment(horizontal='center')

            # Feature rows
            gantt_colors = ['4FC3F7', '81C784', 'FFB74D', 'E57373', 'BA68C8', '4DD0E1', 'F06292', 'AED581']
            for idx, f in enumerate(fa):
                if not isinstance(f, dict): continue
                r = 3 + idx
                ws2.cell(row=r, column=1, value=safe_int(f, "id", idx + 1)).border = thin_border
                ws2.cell(row=r, column=2, value=safe_str(f, "feature")).border = thin_border
                ws2.cell(row=r, column=3, value=safe_str(f, "feature_type")).border = thin_border
                ws2.cell(row=r, column=4, value=safe_int(f, "days")).border = thin_border

                sz = safe_str(f, "size")
                sc = ws2.cell(row=r, column=5, value=sz)
                sc.border = thin_border
                sc.fill = PatternFill("solid", fgColor=size_colors.get(sz, "FFFFFF"))

                ws2.cell(row=r, column=6, value=safe_str(f, "est_team")).border = thin_border
                ws2.cell(row=r, column=7, value=safe_str(f, "est_conservative")).border = thin_border
                ws2.cell(row=r, column=8, value=safe_str(f, "roles_needed")).border = thin_border
                ws2.cell(row=r, column=9, value=safe_str(f, "dependencies", "None")).border = thin_border

                # Gantt bars
                start_d = safe_int(f, "start_day", 1)
                end_d = safe_int(f, "end_day", start_d)
                if start_d < 1: start_d = 1
                if end_d < start_d: end_d = start_d
                color = gantt_colors[idx % len(gantt_colors)]
                for d in range(start_d, min(end_d + 1, total_days + 1)):
                    cell = ws2.cell(row=r, column=num_static + d, value=1)
                    cell.fill = PatternFill("solid", fgColor=color)
                    cell.font = Font(size=7, color=color)
                    cell.alignment = Alignment(horizontal='center')

            # Resource loading row
            r_load_row = 3 + len(fa) + 1
            ws2.cell(row=r_load_row, column=1, value="RESOURCE").font = Font(bold=True, color="FFFFFF")
            ws2.cell(row=r_load_row, column=1).fill = PatternFill("solid", fgColor="EF4444")
            ws2.cell(row=r_load_row, column=2, value="Daily Team Count").font = Font(bold=True)
            for rl in resource_load:
                if not isinstance(rl, dict): continue
                d = safe_int(rl, "day")
                if 1 <= d <= total_days:
                    cell = ws2.cell(row=r_load_row, column=num_static + d, value=safe_int(rl, "team_members_needed"))
                    cell.font = Font(bold=True, size=8)
                    cell.alignment = Alignment(horizontal='center')

            # UAT row
            uat_row = r_load_row + 1
            ws2.cell(row=uat_row, column=1, value="UAT").font = Font(bold=True, color="FFFFFF")
            ws2.cell(row=uat_row, column=1).fill = PatternFill("solid", fgColor="8B5CF6")
            for u in uat:
                if not isinstance(u, dict): continue
                d = safe_int(u, "day")
                if 1 <= d <= total_days:
                    cell = ws2.cell(row=uat_row, column=num_static + d, value=safe_str(u, "name", "UAT"))
                    cell.fill = PatternFill("solid", fgColor="DDD6FE")
                    cell.font = Font(bold=True, size=8, color="6D28D9")

            # Pilot row
            pilot_row = uat_row + 1
            ws2.cell(row=pilot_row, column=1, value="PILOT").font = Font(bold=True, color="FFFFFF")
            ws2.cell(row=pilot_row, column=1).fill = PatternFill("solid", fgColor="F59E0B")
            p_start = safe_int(pilot, "start_day")
            p_end = safe_int(pilot, "end_day")
            for d in range(max(1, p_start), min(p_end + 1, total_days + 1)):
                cell = ws2.cell(row=pilot_row, column=num_static + d, value=1)
                cell.fill = PatternFill("solid", fgColor="FEF3C7")
                cell.font = Font(size=7, color="FEF3C7")

            # Hypercare row
            hc_row = pilot_row + 1
            ws2.cell(row=hc_row, column=1, value="HYPERCARE").font = Font(bold=True, color="FFFFFF")
            ws2.cell(row=hc_row, column=1).fill = PatternFill("solid", fgColor="06B6D4")
            h_start = safe_int(hypercare, "start_day")
            h_end = safe_int(hypercare, "end_day")
            for d in range(max(1, h_start), min(h_end + 1, total_days + 1)):
                cell = ws2.cell(row=hc_row, column=num_static + d, value=1)
                cell.fill = PatternFill("solid", fgColor="CFFAFE")
                cell.font = Font(size=7, color="CFFAFE")

            # Column widths
            ws2.column_dimensions['A'].width = 5
            ws2.column_dimensions['B'].width = 45
            ws2.column_dimensions['C'].width = 18
            ws2.column_dimensions['D'].width = 6
            ws2.column_dimensions['E'].width = 14
            ws2.column_dimensions['F'].width = 22
            ws2.column_dimensions['G'].width = 22
            ws2.column_dimensions['H'].width = 16
            ws2.column_dimensions['I'].width = 18
            for d in range(1, total_days + 1):
                ws2.column_dimensions[get_column_letter(num_static + d)].width = 3.5
            ws2.freeze_panes = ws2.cell(row=3, column=num_static + 1)

            print("[XLSX] Sheet 2 done.", flush=True)
        except Exception as e:
            print(f"[XLSX] Sheet 2 ERROR: {e}", flush=True)
            traceback.print_exc()

        # ═══════════════════════════════════════════════
        # SHEET 3: Gantt Phases
        # ═══════════════════════════════════════════════
        try:
            print("[XLSX] Building Sheet 3: Gantt Phases...", flush=True)
            ws3 = wb.create_sheet("Gantt Phases")
            ws3.sheet_properties.tabColor = "F59E0B"
            write_header(ws3, 1, ["Phase", "Assigned Roles", "Dependencies", "Start Day", "End Day", "Start Week", "End Week", "Duration (days)", "Phase Type"])
            for i, g in enumerate(gantt):
                if not isinstance(g, dict): continue
                r = i + 2
                ws3.cell(row=r, column=1, value=safe_str(g, "phase")).border = thin_border
                ws3.cell(row=r, column=2, value=safe_str(g, "assigned_roles")).border = thin_border
                ws3.cell(row=r, column=3, value=safe_str(g, "dependencies")).border = thin_border
                ws3.cell(row=r, column=4, value=safe_int(g, "start_day")).border = thin_border
                ws3.cell(row=r, column=5, value=safe_int(g, "end_day")).border = thin_border
                ws3.cell(row=r, column=6, value=safe_int(g, "start_week")).border = thin_border
                ws3.cell(row=r, column=7, value=safe_int(g, "end_week")).border = thin_border
                ws3.cell(row=r, column=8, value=safe_int(g, "duration_days")).border = thin_border
                ws3.cell(row=r, column=9, value=safe_str(g, "phase_type")).border = thin_border
            auto_width(ws3)
            print("[XLSX] Sheet 3 done.", flush=True)
        except Exception as e:
            print(f"[XLSX] Sheet 3 ERROR: {e}", flush=True)
            traceback.print_exc()

        # ═══════════════════════════════════════════════
        # SHEET 4: Jira Breakdown
        # ═══════════════════════════════════════════════
        try:
            print("[XLSX] Building Sheet 4: Jira Breakdown...", flush=True)
            ws4 = wb.create_sheet("Jira Breakdown")
            ws4.sheet_properties.tabColor = "8B5CF6"
            write_header(ws4, 1, ["Epic Name", "Feature Type", "Epic Points", "Story Summary", "Description", "Story Points", "Priority"])
            r = 2
            for epic in epics:
                if not isinstance(epic, dict): continue
                stories = safe_list(epic, "stories")
                for story in stories:
                    if not isinstance(story, dict): continue
                    ws4.cell(row=r, column=1, value=safe_str(epic, "epic_name")).border = thin_border
                    ws4.cell(row=r, column=2, value=safe_str(epic, "feature_type")).border = thin_border
                    ws4.cell(row=r, column=3, value=safe_int(epic, "total_points")).border = thin_border
                    ws4.cell(row=r, column=4, value=safe_str(story, "summary")).border = thin_border
                    desc = safe_str(story, "description")
                    ws4.cell(row=r, column=5, value=desc.replace("\n", " | ")[:500]).border = thin_border
                    ws4.cell(row=r, column=6, value=safe_int(story, "story_points")).border = thin_border
                    ws4.cell(row=r, column=7, value=safe_str(story, "priority", "Medium")).border = thin_border
                    r += 1
            auto_width(ws4)
            print("[XLSX] Sheet 4 done.", flush=True)
        except Exception as e:
            print(f"[XLSX] Sheet 4 ERROR: {e}", flush=True)
            traceback.print_exc()

        # ═══════════════════════════════════════════════
        # SHEET 5: Sprint Map
        # ═══════════════════════════════════════════════
        try:
            print("[XLSX] Building Sheet 5: Sprint Map...", flush=True)
            ws5 = wb.create_sheet("Sprint Map")
            ws5.sheet_properties.tabColor = "06B6D4"
            write_header(ws5, 1, ["Sprint", "Start Day", "End Day", "Month", "Calendar Month", "Features", "Points"])
            for i, sm in enumerate(sprint_map):
                if not isinstance(sm, dict): continue
                r = i + 2
                ws5.cell(row=r, column=1, value=safe_str(sm, "sprint")).border = thin_border
                ws5.cell(row=r, column=2, value=safe_int(sm, "start_day")).border = thin_border
                ws5.cell(row=r, column=3, value=safe_int(sm, "end_day")).border = thin_border
                ws5.cell(row=r, column=4, value=safe_str(sm, "month")).border = thin_border
                ws5.cell(row=r, column=5, value=safe_str(sm, "calendar_month")).border = thin_border
                fis = sm.get("features_in_sprint")
                if isinstance(fis, list):
                    features_str = ", ".join(str(x) for x in fis)
                else:
                    features_str = str(fis) if fis else ""
                ws5.cell(row=r, column=6, value=features_str).border = thin_border
                ws5.cell(row=r, column=7, value=safe_int(sm, "points_in_sprint")).border = thin_border
            auto_width(ws5)
            print("[XLSX] Sheet 5 done.", flush=True)
        except Exception as e:
            print(f"[XLSX] Sheet 5 ERROR: {e}", flush=True)
            traceback.print_exc()

        # ═══════════════════════════════════════════════
        # SAVE & RETURN
        # ═══════════════════════════════════════════════
        print("[XLSX] Saving workbook...", flush=True)
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        tech_clean = safe_str(meta, 'tech_stack', 'Project').replace(' ', '_').replace('/', '_')[:20]
        filename = f"Feature_Roadmap_{tech_clean}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        print(f"[XLSX] SUCCESS: {filename} ({output.getbuffer().nbytes} bytes)", flush=True)

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )

    except HTTPException:
        raise
    except Exception as e:
        print(f"[XLSX] FATAL ERROR: {e}", flush=True)
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"XLSX generation failed: {str(e)}")

@app.post("/team/generate_email")
def generate_team_email(payload: dict, creds: dict = Depends(get_jira_creds)):
    """AI-generate a professional email draft for team communication."""
    project = payload.get("project", "")
    recipients = payload.get("recipients", "")
    context = payload.get("context", "")
    
    prompt = f"""You are a professional Scrum Master composing a concise email. 
Project: {project}. Recipients: {recipients}. Context: {context}.
Write a professional, friendly sprint update email. Keep it concise (under 200 words).
Return STRICT JSON: {{"subject": "Clear subject line", "body": "Professional email body with greeting and sign-off"}}"""
    
    try:
        raw = generate_ai_response(prompt, temperature=0.5, force_openai=True)
        if raw:
            parsed = json.loads(raw.replace('```json','').replace('```','').strip())
            return {"subject": parsed.get("subject", ""), "body": parsed.get("body", "")}
    except Exception as e:
        print(f"Email generation error: {e}", flush=True)
    
    return {"subject": f"Sprint Update — {project}", "body": f"Hi Team,\n\nHere is a brief update on project {project}.\n\n[Add your update here]\n\nBest regards"}

@app.post("/team/send_email")
def send_team_email(payload: dict, creds: dict = Depends(get_jira_creds)):
    """Send email directly via SMTP. Credentials come from server environment variables — customers never see them."""
    import smtplib
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart
    
    # SMTP credentials from environment (admin-configured, not customer-facing)
    smtp_host = os.getenv("SMTP_HOST", "smtp.gmail.com")
    smtp_port = int(os.getenv("SMTP_PORT", "587"))
    smtp_email = os.getenv("SMTP_EMAIL", "")
    smtp_password = os.getenv("SMTP_PASSWORD", "")
    sender_name = os.getenv("SMTP_SENDER_NAME", "IG Agile Scrum")
    
    recipients = payload.get("recipients", [])
    subject = payload.get("subject", "")
    body = payload.get("body", "")
    project = payload.get("project", "")
    
    if not smtp_email or not smtp_password:
        return {"status": "error", "message": "Email service not configured. Please ask your administrator to set SMTP_EMAIL and SMTP_PASSWORD environment variables."}
    if not recipients:
        return {"status": "error", "message": "No recipients specified."}
    if not subject and not body:
        return {"status": "error", "message": "Email subject and body are both empty."}
    
    try:
        msg = MIMEMultipart('alternative')
        msg['From'] = f"{sender_name} <{smtp_email}>"
        msg['To'] = ', '.join(recipients)
        msg['Subject'] = subject
        msg['Reply-To'] = smtp_email
        
        # Plain text version
        msg.attach(MIMEText(body, 'plain'))
        
        # HTML version with professional formatting
        html_body = f"""
        <div style="font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; max-width: 600px; margin: 0 auto; color: #1a1a2e;">
            <div style="border-bottom: 3px solid #3B82F6; padding-bottom: 16px; margin-bottom: 24px;">
                <strong style="color: #3B82F6; font-size: 13px; letter-spacing: 1px;">IG AGILE SCRUM</strong>
            </div>
            {''.join(f'<p style="margin: 0 0 12px 0; line-height: 1.7; font-size: 15px;">{line}</p>' for line in body.split(chr(10)) if line.strip())}
            <div style="border-top: 1px solid #e5e7eb; margin-top: 32px; padding-top: 16px;">
                <small style="color: #6b7280; font-size: 11px;">Sent via IG Agile Scrum — Enterprise Agile Intelligence Platform</small>
            </div>
        </div>"""
        msg.attach(MIMEText(html_body, 'html'))
        
        # Connect and send
        print(f"[EMAIL] Connecting to {smtp_host}:{smtp_port}...", flush=True)
        with smtplib.SMTP(smtp_host, smtp_port, timeout=15) as server:
            server.ehlo()
            server.starttls()
            server.ehlo()
            server.login(smtp_email, smtp_password)
            server.sendmail(smtp_email, recipients, msg.as_string())
        
        print(f"[EMAIL] ✅ Sent to {len(recipients)} recipients: {', '.join(recipients)}", flush=True)
        return {"status": "sent", "recipients_count": len(recipients)}
    
    except smtplib.SMTPAuthenticationError as e:
        print(f"[EMAIL] ❌ Auth failed: {e}", flush=True)
        return {"status": "error", "message": "Email authentication failed. Please contact your administrator to verify SMTP credentials."}
    except smtplib.SMTPRecipientsRefused as e:
        print(f"[EMAIL] ❌ Recipients refused: {e}", flush=True)
        return {"status": "error", "message": "One or more recipient addresses were rejected by the mail server."}
    except smtplib.SMTPException as e:
        print(f"[EMAIL] ❌ SMTP Error: {e}", flush=True)
        return {"status": "error", "message": f"Mail server error: {str(e)}"}
    except Exception as e:
        print(f"[EMAIL] ❌ General Error: {e}", flush=True)
        return {"status": "error", "message": f"Failed to send: {str(e)}"}

def process_silent_webhook(issue_key, summary, desc_text, project_key, creds_dict):
    try:
        print(f"[1/6] Silent Agent started for: {issue_key}", flush=True)
        time.sleep(3)
        sp_field = get_story_point_field(creds_dict)
        print(f"[2/6] Fetching robust Omni-Roster...", flush=True)
        roster, assignable_map = build_team_roster(project_key, creds_dict, sp_field)
        prompt = f"You are an Autonomous Scrum Master. Ticket: Summary: {summary} | Description: {desc_text}. Roster (MUST pick EXACT NAME from keys): {json.dumps(roster)}. Tasks: 1. Assign Points. 2. Choose Assignee. 3. If Description is short, rewrite it. Return STRICT JSON OBJECT ONLY: {{\"points\": 3, \"assignee\": \"Exact Name\", \"generated_description\": \"Full description\", \"reasoning\": \"Explanation\"}}"
        print(f"[3/6] Querying AI...", flush=True)
        raw = generate_ai_response(prompt, temperature=0.4, force_openai=True)
        if not raw: return
        est = json.loads(raw.replace('```json','').replace('```','').strip())
        target_assignee = est.get('assignee', '')
        assignee_id = assignable_map.get(target_assignee)
        update_fields_basic = {}
        if assignee_id: update_fields_basic["assignee"] = {"accountId": assignee_id}
        gen_desc = est.get("generated_description", "")
        if gen_desc and len(desc_text.strip()) < 20: update_fields_basic["description"] = create_adf_doc("AI Generated Description:\n\n" + gen_desc)
        if update_fields_basic:
            print(f"[5a/6] Updating Description & Assignee...", flush=True)
            jira_request("PUT", f"issue/{issue_key}", creds_dict, {"fields": update_fields_basic})
        points = safe_float(est.get('points', 0))
        if points > 0:
            print(f"[5b/6] Updating Story Points ({points})...", flush=True)
            jira_request("PUT", f"issue/{issue_key}", creds_dict, {"fields": {sp_field: points}})
        print(f"[6/6] Posting Insight Comment to Jira...", flush=True)
        comment_text = f"IG Agile Auto-Triage Complete\n- Estimated Points: {points}\n- Suggested Assignee: {target_assignee}\n- Reasoning: {est.get('reasoning', '')}\n"
        if gen_desc and len(desc_text.strip()) < 20: comment_text += f"\n\nGenerated Description:\n{gen_desc}"
        jira_request("POST", f"issue/{issue_key}/comment", creds_dict, {"body": create_adf_doc(comment_text)})
        print(f"Webhook Process Complete for {issue_key}", flush=True)
    except Exception as e: print(f"FATAL Webhook Exception for {issue_key}: {e}", flush=True)

@app.post("/webhook")
async def jira_webhook(request: Request, background_tasks: BackgroundTasks, domain: str = None, email: str = None, token: str = None, cloud_id: str = None):
    try:
        payload = await request.json()
        if payload.get("webhookEvent") != "jira:issue_created": return {"status": "ignored"}
        issue = payload.get("issue", {}); key = issue.get("key"); fields = issue.get("fields", {})
        desc_raw = fields.get("description")
        desc = ""
        if desc_raw:
            if isinstance(desc_raw, str): desc = desc_raw[:500]
            else: desc = extract_adf_text(desc_raw)[:500]
        summary = fields.get("summary", "")
        project_key = (fields.get("project") or {}).get("key", "")
        if "IG Agile AI Insights" in desc or "AI Generated Description" in desc or "AI Generated Story" in desc:
            print(f"Skipping Webhook for {key}: Issue was created actively by the UI.", flush=True)
            return {"status": "ignored"}
        print(f"\nWEBHOOK FIRED: New Issue {key} detected in project {project_key}.", flush=True)
        creds_dict = None
        if domain and email and token: creds_dict = {"auth_type": "basic", "domain": domain, "email": email, "token": token}
        else:
            db = SessionLocal()
            if cloud_id: user = db.query(UserAuth).filter(UserAuth.cloud_id == cloud_id).first()
            else: user = db.query(UserAuth).order_by(UserAuth.expires_at.desc()).first()
            if user:
                fresh_user = get_valid_oauth_session(db=db, license_key=user.license_key)
                if fresh_user: creds_dict = {"auth_type": "oauth", "cloud_id": fresh_user.cloud_id, "access_token": fresh_user.access_token}
            db.close()
        if creds_dict: background_tasks.add_task(process_silent_webhook, key, summary, desc, project_key, creds_dict)
        return {"status": "processing_in_background"}
    except Exception as e: return {"status": "error", "message": str(e)}

@app.post("/roadmap/{project_key}/investment_buckets")
def update_investment_buckets(
    project_key: str,
    payload: dict,
    creds: dict = Depends(get_jira_creds),
):
    '''
    Update investment bucket ratios for the project.
    Leadership decides: 60% Features, 20% Tech Debt, 15% Bugs, 5% Innovation.
    '''
    buckets = payload.get("buckets", {})
    
    # Validate percentages sum to 100
    total_pct = sum(b.get("pct", 0) for b in buckets.values())
    if abs(total_pct - 100) > 1:
        raise HTTPException(status_code=400, detail=f"Bucket percentages must sum to 100 (got {total_pct})")
    
    # Store in Jira project properties
    jira_request(
        "PUT",
        f"project/{project_key.upper()}/properties/ig_agile_investment_buckets",
        creds,
        {"buckets": buckets, "updated_at": datetime.utcnow().isoformat()}
    )
    
    return {"status": "saved", "buckets": buckets}



# ================= MEETING AGENT ENDPOINTS =================

@app.post("/meeting/upload")
async def meeting_upload(payload: dict, background_tasks: BackgroundTasks, creds: dict = Depends(get_jira_creds), db: Session = Depends(get_db)):
    """Upload/paste a meeting transcript for AI processing."""
    transcript = payload.get("transcript", "")
    project_key = payload.get("project", "")
    sprint_id = payload.get("sprint_id") or payload.get("sprint", "")
    meeting_type = payload.get("meeting_type")  # Optional — auto-detect if not provided
    platform = payload.get("platform", "manual")

    if not transcript or len(transcript.strip()) < 50:
        raise HTTPException(status_code=400, detail="Transcript is too short. Please provide at least 50 characters.")
    if not project_key:
        raise HTTPException(status_code=400, detail="Project key is required.")

    session_id = str(uuid.uuid4())
    session = MeetingSession(
        id=session_id, license_key=payload.get("license_key", ""),
        project_key=project_key, sprint_id=str(sprint_id) if sprint_id else "",
        meeting_type=meeting_type or "auto", transcript=transcript,
        status="processing", platform=platform
    )
    db.add(session); db.commit()

    # Get team roster for smart assignment
    sp_field = get_story_point_field(creds)
    roster, assignable_map = build_team_roster(project_key, creds, sp_field)
    team_roster = roster  # {name: points} dict

    # Process transcript (this can take 30-60s, so we do it synchronously for now
    # since the frontend shows a loading state)
    try:
        results = process_meeting_transcript(
            transcript=transcript,
            project_key=project_key,
            sprint_id=sprint_id,
            meeting_type=meeting_type,
            jira_request_fn=jira_request,
            creds=creds,
            team_roster=team_roster
        )

        # Update session with results
        session = db.query(MeetingSession).filter(MeetingSession.id == session_id).first()
        if session:
            session.status = "completed"
            session.meeting_type = results.get("meeting_type", meeting_type or "unknown")
            session.ai_results = json.dumps(results, default=str)
            db.commit()

        return {"status": "success", "session_id": session_id, "results": results}

    except Exception as e:
        print(f"Meeting Agent Error: {e}", flush=True)
        traceback.print_exc()
        session = db.query(MeetingSession).filter(MeetingSession.id == session_id).first()
        if session:
            session.status = "error"
            session.ai_results = json.dumps({"error": str(e)})
            db.commit()
        return {"status": "error", "session_id": session_id, "message": str(e)}


@app.get("/meeting/sessions/{project_key}")
def list_meeting_sessions(project_key: str, creds: dict = Depends(get_jira_creds), db: Session = Depends(get_db)):
    """List past meeting sessions for a project."""
    sessions = db.query(MeetingSession).filter(
        MeetingSession.project_key == project_key
    ).order_by(MeetingSession.created_at.desc()).limit(20).all()
    return [{"id": s.id, "meeting_type": s.meeting_type, "status": s.status,
             "platform": s.platform, "created_at": s.created_at.isoformat() if s.created_at else "",
             "transcript_preview": (s.transcript or "")[:200]} for s in sessions]


@app.get("/meeting/session/{session_id}")
def get_meeting_session(session_id: str, creds: dict = Depends(get_jira_creds), db: Session = Depends(get_db)):
    """Get full session details and AI results."""
    session = db.query(MeetingSession).filter(MeetingSession.id == session_id).first()
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    results = None
    try:
        results = json.loads(session.ai_results) if session.ai_results else None
    except: pass
    return {
        "id": session.id, "project_key": session.project_key, "sprint_id": session.sprint_id,
        "meeting_type": session.meeting_type, "status": session.status,
        "platform": session.platform, "transcript": session.transcript,
        "results": results, "created_at": session.created_at.isoformat() if session.created_at else ""
    }


@app.post("/meeting/session/{session_id}/approve")
async def approve_meeting_stories(session_id: str, payload: dict, creds: dict = Depends(get_jira_creds), db: Session = Depends(get_db)):
    """Approve AI-generated stories and create them in Jira."""
    session = db.query(MeetingSession).filter(MeetingSession.id == session_id).first()
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")

    try:
        results = json.loads(session.ai_results) if session.ai_results else {}
    except:
        raise HTTPException(status_code=400, detail="No AI results available")

    story_indices = payload.get("story_indices", [])  # Which stories to create
    epic_indices = payload.get("epic_indices", [])     # Which epics to create
    auto_assign = payload.get("auto_assign", True)
    stories = results.get("extracted", {}).get("stories", [])
    epics = results.get("extracted", {}).get("epics", [])
    project_key = session.project_key
    sp_field = get_story_point_field(creds)
    roster, assignable_map = build_team_roster(project_key, creds, sp_field)

    # Get base URL for issue links
    base_url = ""
    if creds.get("auth_type") == "basic":
        base_url = f"https://{creds['domain']}"
    else:
        server_info = jira_request("GET", "serverInfo", creds)
        if server_info and server_info.status_code == 200:
            base_url = server_info.json().get("baseUrl", "")

    created_issues = []
    errors = []

    # Create epics first
    epic_key_map = {}  # index -> jira key
    for idx in epic_indices:
        if idx < len(epics):
            epic = epics[idx]
            desc_text = f"Motivation:\n{epic.get('motivation', '')}\n\nDescription:\n{epic.get('description', '')}"
            issue_data = {
                "fields": {
                    "project": {"key": project_key},
                    "summary": epic.get("title", "AI Generated Epic"),
                    "description": create_adf_doc(desc_text),
                    "issuetype": {"name": "Epic"}
                }
            }
            res = jira_request("POST", "issue", creds, issue_data)
            if res and res.status_code == 201:
                ek = res.json().get("key")
                epic_key_map[idx] = ek
                url = f"{base_url}/browse/{ek}" if base_url else ""
                created_issues.append({"type": "Epic", "key": ek, "title": epic.get("title"), "url": url})
            else:
                errors.append(f"Failed to create epic: {epic.get('title')}")

    # Create stories
    for idx in story_indices:
        if idx < len(stories):
            story = stories[idx]
            desc_text = story.get("description", "AI Generated Story")
            ac_list = story.get("acceptance_criteria", [])
            issue_data = {
                "fields": {
                    "project": {"key": project_key},
                    "summary": story.get("title", "AI Generated Story"),
                    "description": create_adf_doc(desc_text, ac_list),
                    "issuetype": {"name": "Story"}
                }
            }
            res = jira_request("POST", "issue", creds, issue_data)
            if res is None or res.status_code != 201:
                # Fallback to Task
                issue_data["fields"]["issuetype"]["name"] = "Task"
                res = jira_request("POST", "issue", creds, issue_data)

            if res and res.status_code == 201:
                new_key = res.json().get("key")

                # Set story points
                points = safe_float(story.get("suggested_points", 0))
                if points > 0:
                    jira_request("PUT", f"issue/{new_key}", creds, {"fields": {sp_field: points}})

                # Assign
                if auto_assign and story.get("suggested_assignee") and story["suggested_assignee"].lower() != "unassigned":
                    aid = assignable_map.get(story["suggested_assignee"])
                    if not aid:
                        aid = get_jira_account_id(story["suggested_assignee"], creds)
                    if aid:
                        jira_request("PUT", f"issue/{new_key}", creds, {"fields": {"assignee": {"accountId": aid}}})

                # Add AI insights comment
                comment = (f"🤖 IG Agile Meeting Agent — Auto-Generated\n"
                           f"- Estimated: {points} pts\n"
                           f"- Sprint Fit: {story.get('sprint_fit', 'unknown')}\n"
                           f"- Reasoning: {story.get('estimation_reasoning', '')}\n"
                           f"- Capacity: {story.get('capacity_reasoning', '')}")
                jira_request("POST", f"issue/{new_key}/comment", creds,
                             {"body": create_adf_doc(comment)})

                url = f"{base_url}/browse/{new_key}" if base_url else ""
                created_issues.append({"type": "Story", "key": new_key, "title": story.get("title"),
                                        "points": points, "assignee": story.get("suggested_assignee"), "url": url})
            else:
                errors.append(f"Failed to create: {story.get('title')} — {extract_jira_error(res)}")

    return {"status": "success", "created": created_issues, "errors": errors,
            "total_created": len(created_issues)}


@app.get("/meeting/history/{project_key}")
def get_sprint_history(project_key: str, num_sprints: int = 5, creds: dict = Depends(get_jira_creds)):
    """Get sprint history and velocity data for capacity planning."""
    history = fetch_sprint_history(jira_request, creds, project_key, num_sprints)
    velocity = calculate_velocity(history)
    return {"sprint_history": [{"name": h["sprint"]["name"], "velocity": h["completed_points"],
             "completion_rate": h["completion_rate"], "total_issues": h["total_issues"],
             "completed_issues": h["completed_issues"],
             "person_stats": h["person_stats"]} for h in history],
            "velocity": velocity}


@app.get("/meeting/capacity/{project_key}")
def get_capacity_analysis(project_key: str, sprint_id: str = None, creds: dict = Depends(get_jira_creds)):
    """Full capacity analysis for current sprint."""
    sp_field = get_story_point_field(creds)
    history = fetch_sprint_history(jira_request, creds, project_key)
    velocity = calculate_velocity(history)

    # Get current sprint issues
    jql = f'project="{project_key}" AND sprint={sprint_id}' if sprint_id and sprint_id != "active" else f'project="{project_key}" AND sprint in openSprints()'
    res = jira_request("POST", "search/jql", creds, {
        "jql": jql, "maxResults": 100,
        "fields": ["summary", "status", "assignee", sp_field, "customfield_10016", "customfield_10026"]
    })
    current_issues = []
    if res and res.status_code == 200:
        for iss in res.json().get('issues', []):
            f = iss.get('fields') or {}
            pts = extract_story_points(f, sp_field)
            status_name = (f.get('status') or {}).get('name', 'To Do')
            assignee_name = (f.get('assignee') or {}).get('displayName', 'Unassigned')
            current_issues.append({"key": iss.get('key'), "summary": f.get('summary', ''),
                                    "assignee": assignee_name, "points": pts, "status": status_name})

    report = generate_capacity_report(history, velocity, current_issues)
    report["current_sprint_issues"] = current_issues
    return report


@app.post("/meeting/webhook/transcript")
async def meeting_transcript_webhook(request: Request, background_tasks: BackgroundTasks, db: Session = Depends(get_db)):
    """Webhook receiver for Recall.ai / MeetingBaaS transcript delivery."""
    try:
        payload = await request.json()
        transcript = payload.get("transcript", "")
        meeting_url = payload.get("meeting_url", "")
        platform = payload.get("platform", "recall_ai")
        project_key = payload.get("project_key", "")
        sprint_id = payload.get("sprint_id", "")
        license_key = payload.get("license_key", "")

        if not transcript or not project_key:
            return {"status": "error", "message": "Missing transcript or project_key"}

        session_id = str(uuid.uuid4())
        session = MeetingSession(
            id=session_id, license_key=license_key,
            project_key=project_key, sprint_id=sprint_id,
            meeting_type="auto", transcript=transcript,
            status="received", platform=platform
        )
        db.add(session); db.commit()

        # Process in background
        creds_dict = None
        if license_key:
            user = get_valid_oauth_session(db=db, license_key=license_key)
            if user:
                creds_dict = {"auth_type": "oauth", "cloud_id": user.cloud_id, "access_token": user.access_token}

        if creds_dict:
            background_tasks.add_task(
                _process_webhook_transcript, session_id, transcript,
                project_key, sprint_id, creds_dict
            )

        return {"status": "accepted", "session_id": session_id}
    except Exception as e:
        return {"status": "error", "message": str(e)}




def _process_webhook_transcript(session_id, transcript, project_key, sprint_id, creds):
    """Background task to process a webhook-delivered transcript."""
    try:
        sp_field = get_story_point_field(creds)
        roster, _ = build_team_roster(project_key, creds, sp_field)

        results = process_meeting_transcript(
            transcript=transcript, project_key=project_key,
            sprint_id=sprint_id, jira_request_fn=jira_request,
            creds=creds, team_roster=roster
        )

        db = SessionLocal()
        session = db.query(MeetingSession).filter(MeetingSession.id == session_id).first()
        if session:
            session.status = "completed"
            session.meeting_type = results.get("meeting_type", "unknown")
            session.ai_results = json.dumps(results, default=str)
            db.commit()
        db.close()
        print(f"Webhook transcript processed: {session_id}", flush=True)
    except Exception as e:
        print(f"Webhook transcript error: {e}", flush=True)
        db = SessionLocal()
        session = db.query(MeetingSession).filter(MeetingSession.id == session_id).first()
        if session:
            session.status = "error"
            session.ai_results = json.dumps({"error": str(e)})
            db.commit()
        db.close()


@app.post("/meeting/analyze_capacity")
async def analyze_capacity_on_demand(payload: dict, creds: dict = Depends(get_jira_creds)):
    """On-demand capacity planning from sprint data (no transcript needed)."""
    project_key = payload.get("project", "")
    sprint_id = payload.get("sprint_id", "")
    if not project_key:
        raise HTTPException(status_code=400, detail="Project key required")

    sp_field = get_story_point_field(creds)
    history = fetch_sprint_history(jira_request, creds, project_key)
    velocity = calculate_velocity(history)

    # Current sprint load
    jql = f'project="{project_key}" AND sprint={sprint_id}' if sprint_id and sprint_id != "active" else f'project="{project_key}" AND sprint in openSprints()'
    res = jira_request("POST", "search/jql", creds, {
        "jql": jql, "maxResults": 100,
        "fields": ["summary", "status", "assignee", sp_field, "customfield_10016", "customfield_10026"]
    })
    current_issues = []
    if res and res.status_code == 200:
        for iss in res.json().get('issues', []):
            f = iss.get('fields') or {}
            pts = extract_story_points(f, sp_field)
            status_name = (f.get('status') or {}).get('name', 'To Do')
            assignee_name = (f.get('assignee') or {}).get('displayName', 'Unassigned')
            current_issues.append({"key": iss.get('key'), "summary": f.get('summary', ''),
                                    "assignee": assignee_name, "points": pts, "status": status_name})

    report = generate_capacity_report(history, velocity, current_issues)

    # AI-powered recommendations
    prompt = f"""You are an Agile capacity planning expert. Analyze this data and provide recommendations.

Team Velocity (last {velocity.get('sprints_analyzed', 0)} sprints): {json.dumps(velocity)}
Current Sprint Issues: {json.dumps(current_issues[:20])}
Team Capacity: {json.dumps(report.get('team_capacity', {}))}

Return JSON: {{"capacity_summary": "2-3 sentence executive summary", "risk_level": "low|medium|high", "recommendations": ["Actionable recommendation 1", "Rec 2"], "team_health": "Brief assessment"}}"""

    try:
        raw = generate_ai_response(prompt, temperature=0.3)
        ai_analysis = json.loads(raw.replace('```json', '').replace('```', '').strip()) if raw else {}
    except: ai_analysis = {}

    report["ai_analysis"] = ai_analysis
    report["current_sprint_issues"] = current_issues
    return report